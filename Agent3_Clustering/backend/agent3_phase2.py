"""
Agent 3 Phase 2 — Cross-MH DH reassignment optimizer.

Takes Agent 3's output (dh_fc_mh_assignment.csv), the H2H network file, and
all Agent 4 inputs.  For every MH pair where Agent 3 flagged DHs for
reassignment, the pool is expanded via H2H MR-group logic, the best split
between the two MHs is found using Agent 4's full routing pipeline as the
scoring function, then before/after Excel reports are produced.

Issue-fix log
-------------
I1  Enumeration and LNS now call _run_a4_subset (full ILP) for every candidate
    assignment — no simplified proxy anywhere in the winner-selection path.
I2  MH-MH delta for route-mates (not in Agent 3 candidate columns) resolved via
    smh_mhlast_cost_per_shipment.csv; flagged None + note when unavailable.
I3  MHDH_Before sheet shows one row per H2H MR group (current routes), not one
    row per DH.
I4  Non-pool costs are precomputed once and added as a fixed offset to all
    reported costs so totals reflect the full MH, not just the contested pool.
I5  LNS repair uses _dh_direct_cost for greedy assignment; the resulting
    candidate is always evaluated with _run_a4_subset (full ILP).
I6  Baseline already uses _run_a4_subset — confirmed correct, no change needed.
I7  Per_DH_Detail gains Saving_confirmed (per-DH delta > 3 000 Rs).
    Summary gains Pair_saving_confirmed (total pair saving > 3 000 Rs).
I8  City/MH filter in Streamlit UI comes from Agent 3 output (see tab file).
I9  MHDH_Before now uses before_from_result/before_to_result.final_assignment_df
    (Agent 4 baseline ILP) instead of H2H proxy — exact costs, zero NaN.
I10 Fixed NameError: flagged_orig used consistently in MHPairResult constructor.
I11 _mhmh_cost_at_mh now reads candidate_X_mhmh_cost_rs (exact MH-MH, not blended
    total). Gap-fill cache from compute_mhmh_for_pairs() checked as Step 2.
    SMH proxy fallback removed — if both steps miss, cost is "unavailable" (honest).
    run_phase2_pipeline() accepts plan_vol_path + mh1_mh2_cost_matrix_path (from
    orchestrator) and batch-fills missing MH-MH costs before each pair's Excel build.
    Per-DH MH-MH source logged to progress output for full visibility.
I12 Non-pool cost precomputation removed entirely. cost_before = pool_cost_before,
    cost_after = pool_cost_after. Non_Pool_Cost_Rs column removed from all outputs.
I13 Summary redesigned: Pool_MHDH_Cost_Delta_Rs + Pool_MHMH_Cost_Delta_Rs + Total_Savings_Rs.
    Per_DH_Detail redesigned: MHDH costs from ILP route (route_cost/n_dhs, not solo proxy),
    MHMH costs on both current and target sides (×30 for monthly), Total_Saving per DH.
    MHMH current side now resolved via candidate columns for from_mh (same logic as target
    side) — fixes NaN for DHs where Agent 3 assigned them to to_mh. No gap-fill needed for
    current side if from_mh is in candidates. The two _run_a4_subset() calls on
    all non-pool DHs at each MH were running full ILP on 30-60 DHs per pair and adding
    minutes of wasted computation. Non-pool cost cancels out in savings (before−after),
    so removing it has zero effect on the savings number. cost_before = pool_cost_before,
    cost_after = pool_cost_after. Non_Pool_Cost_Rs column removed from all outputs.

Public API
----------
    load_h2h(path)                          -> pd.DataFrame
    compute_mh_pair_savings(agent3_df)      -> pd.DataFrame
    expand_pool(...)                        -> list[str]
    optimize_pool_assignment(...)           -> tuple[dict, A4Result, A4Result]
    build_pair_output(..., mhmh_cache)      -> dict[str, pd.DataFrame]
    run_phase2_pipeline(                    -> Phase2Result
        ...,
        plan_vol_path,               # from orchestrator — enables MH-MH gap-fill
        mh1_mh2_cost_matrix_path,    # from orchestrator — enables MH-MH gap-fill
    )
    write_excel_outputs(...)               -> list[Path]
"""
from __future__ import annotations

import dataclasses
import random
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Optional

import pandas as pd

# ── Agent 4 backend on path ──────────────────────────────────────────────────
_A4_BACKEND = Path(r"C:\Users\aniket.kathuria\Desktop\Agentic tools\Agent4_Routing\backend")
if _A4_BACKEND.is_dir() and str(_A4_BACKEND) not in sys.path:
    sys.path.insert(0, str(_A4_BACKEND))

import agent4 as p4
from agent3_pipeline import compute_mhmh_for_pairs as _a3_compute_mhmh

# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

HUB_PREFIXES = (
    "SATELLITEHUB", "CENTRALHUB", "BULKHUB",
    "LINEHAULHUB", "FC_MH", "LM_",
)
DAYS_PER_MONTH      = 30
LNS_ITERATIONS      = 500
FULL_ENUM_THRESHOLD = 15
SAVING_CONFIRMED_THRESHOLD_RS = 3_000.0


# ─────────────────────────────────────────────────────────────────────────────
# Result dataclasses
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class MHPairResult:
    """Outcome of Phase 2 optimization for one (from_mh → to_mh) pair."""
    from_mh:            str
    to_mh:              str
    flagged_dhs:        list[str]
    pool_dhs:           list[str]
    best_assignment:    dict[str, str]          # dh → assigned mh_name
    # Pool costs (from Agent 4 runs on pool DHs only)
    pool_cost_before:   float
    pool_cost_after:    float
    cost_before:        float                   # = pool_cost_before
    cost_after:         float                   # = pool_cost_after
    savings:            float
    dhs_moved:          list[str]
    dhs_stayed:         list[str]
    # MHMH monthly delta (current − target) × 30, summed across pool DHs where available
    mhmh_delta_rs:           Optional[float] = None
    # Agent 3 estimated monthly saving for DHs that Phase 2 actually moved (cost_delta_rs × 30)
    agent3_est_saving_rs:    Optional[float] = None
    # Agent 4 results (before/after, each MH, pool DHs only)
    before_from_result: Optional[p4.Agent4MHResult] = None
    before_to_result:   Optional[p4.Agent4MHResult] = None
    after_from_result:  Optional[p4.Agent4MHResult] = None
    after_to_result:    Optional[p4.Agent4MHResult] = None
    sheets:   dict[str, pd.DataFrame] = field(default_factory=dict)
    map_html: Optional[str] = None   # self-contained Leaflet HTML for the pool map


@dataclass
class Phase2Result:
    """Full Phase 2 pipeline output."""
    pair_results: list[MHPairResult]
    summary_df:   pd.DataFrame
    excel_paths:  list[Path]


# ─────────────────────────────────────────────────────────────────────────────
# Phase 2 helper: Load H2H
# ─────────────────────────────────────────────────────────────────────────────

def load_h2h(path: Path) -> pd.DataFrame:
    """
    Load and normalise the H2H network file.

    Added columns
    -------------
    Src_upper    : Src stripped + upper-cased
    Dest_upper   : Dest stripped + upper-cased
    MR_cleaned   : blank / NaN / #REF! / "Direct" → ``DIRECT_{Dest_upper}``
                   all other values kept as-is
    grouping_key : tuple (Src_upper, MR_cleaned)

    Only rows whose Src_upper starts with a known hub prefix are kept.
    """
    df = pd.read_csv(path, dtype=str)
    df.columns = df.columns.str.strip()

    df["Src_upper"]  = df["Src"].str.strip().str.upper().fillna("")
    df["Dest_upper"] = df["Dest"].str.strip().str.upper().fillna("")

    def _clean_mr(row: pd.Series) -> str:
        dest = row["Dest_upper"]
        raw  = str(row.get("MR Number", "")).strip()
        if raw.lower() in ("", "nan", "#ref!", "direct"):
            return f"DIRECT_{dest}"
        return raw

    df["MR_cleaned"]   = df.apply(_clean_mr, axis=1)
    df["grouping_key"] = list(zip(df["Src_upper"], df["MR_cleaned"]))

    mask = df["Src_upper"].str.startswith(
        tuple(p.upper() for p in HUB_PREFIXES), na=False
    )
    return df[mask].reset_index(drop=True)


# ─────────────────────────────────────────────────────────────────────────────
# Phase 3: Compute MH-pair savings from Agent 3 output
# ─────────────────────────────────────────────────────────────────────────────

def compute_mh_pair_savings(agent3_df: pd.DataFrame) -> pd.DataFrame:
    """
    For each DH where assigned_fc_mh != current_fc_mh, compute per-DH savings
    and aggregate to MH-pair level.

    Returns DataFrame with columns:
    from_mh, to_mh, n_dhs_flagged, total_savings_rs, dhs (list of DH keys)
    sorted by total_savings_rs descending.
    """
    df = agent3_df.copy()
    moved = df[
        df["assigned_fc_mh"].astype(str).str.strip()
        != df["current_fc_mh"].astype(str).str.strip()
    ].copy()

    moved["dh_savings_rs"] = (
        pd.to_numeric(moved["current_fc_cost_rs"], errors="coerce").fillna(0)
        - pd.to_numeric(moved["total_cost_rs"],    errors="coerce").fillna(0)
    )

    pairs: list[dict] = []
    for (from_mh, to_mh), grp in moved.groupby(
        [moved["current_fc_mh"].str.strip(), moved["assigned_fc_mh"].str.strip()]
    ):
        pairs.append({
            "from_mh":          from_mh,
            "to_mh":            to_mh,
            "n_dhs_flagged":    len(grp),
            "total_savings_rs": round(grp["dh_savings_rs"].sum(), 2),
            "dhs":              grp["destination_hub_key"].tolist(),
        })

    return (
        pd.DataFrame(pairs)
        .sort_values("total_savings_rs", ascending=False)
        .reset_index(drop=True)
    )


# ─────────────────────────────────────────────────────────────────────────────
# Phase 4: Expand pool via H2H MR-group logic
# ─────────────────────────────────────────────────────────────────────────────

def expand_pool(
    flagged_dhs:  list[str],
    current_mh:   str,
    target_mh:    str,
    h2h_df:       pd.DataFrame,
    agent3_df:    pd.DataFrame,
    location_df:  pd.DataFrame,
) -> list[str]:
    """
    Expand the flagged DH pool using H2H MR-group logic.

    Rules
    -----
    - Only use H2H rows where Src_upper == current_mh.upper()
    - DIRECT_* MR codes never expand the pool
    - For each non-DIRECT MR group touching a flagged DH, add all other DHs
      in the same (Src_upper, MR_cleaned) group that exist in location_df AND
      are currently served by either current_mh or target_mh
    - Cross-MH MR codes: use only rows where Src_upper == current_mh.upper()

    Returns deduplicated, sorted list of DH keys (original casing from
    location_df where available).
    """
    mh_upper     = current_mh.strip().upper()
    target_upper = target_mh.strip().upper()

    mh_h2h = h2h_df[h2h_df["Src_upper"] == mh_upper]

    loc_mh_col  = location_df["current_fc_mh"].str.strip().str.upper()
    valid_upper = set(
        location_df[loc_mh_col.isin({mh_upper, target_upper})][
            "destination_hub_key"
        ].str.strip().str.upper()
    )

    mr_to_dhs: dict[str, set[str]] = {}
    for _, row in mh_h2h.iterrows():
        mr = row["MR_cleaned"]
        if mr.startswith("DIRECT_"):
            continue
        mr_to_dhs.setdefault(mr, set()).add(row["Dest_upper"])

    flagged_upper = {d.strip().upper() for d in flagged_dhs}
    pool: set[str] = set(flagged_upper)

    for mr, dh_set in mr_to_dhs.items():
        if dh_set & flagged_upper:
            pool |= dh_set & valid_upper

    orig_map = {
        k.strip().upper(): k
        for k in location_df["destination_hub_key"].tolist()
    }
    return sorted(orig_map.get(d, d) for d in pool)


# ─────────────────────────────────────────────────────────────────────────────
# Agent 4 runner for a DH subset
# ─────────────────────────────────────────────────────────────────────────────

def _empty_mh_result(mh_name: str) -> p4.Agent4MHResult:
    return p4.Agent4MHResult(
        mh_name=mh_name,
        clustering_df=pd.DataFrame(),
        filtered_routes_df=pd.DataFrame(),
        final_assignment_df=pd.DataFrame(),
        expanded_schedule_df=pd.DataFrame(),
        validation_lines=[],
        total_monthly_cost=0.0,
        n_clusters=0,
        n_perms_checked=0,
        n_routes_survived=0,
        ilp_status={},
        missing_dhs=[],
        absorbed_residuals_df=pd.DataFrame(),
        dh_summary_df=pd.DataFrame(),
    )


def _run_a4_subset(
    mh_name:            str,
    dh_keys:            list[str],
    dist_dict:          dict,
    latlong:            dict,
    mh_configs:         dict,
    location_df:        pd.DataFrame,
    cfg:                dict,
    residual_threshold: float = 100.0,
) -> p4.Agent4MHResult:
    """Run Agent 4's full routing pipeline (Steps 2-4) for a DH subset at mh_name."""
    if not dh_keys:
        return _empty_mh_result(mh_name)

    mh_cfg = mh_configs.get(mh_name)
    if mh_cfg is None:
        mh_cfg = p4.MHConfig(
            mh_name=mh_name,
            local_rate_card={6.5: 12.0, 8: 16.0, 10: 20.0, 14: 24.0,
                             17: 28.0, 20: 32.0, 22: 36.0, 24: 40.0,
                             32: 48.0, 40: 56.0},
            zonal_rate_card={6.5: 10.0, 8: 13.0, 10: 17.0, 14: 21.0,
                             17: 25.0, 20: 29.0, 22: 33.0, 24: 37.0,
                             32: 45.0, 40: 53.0},
            max_hops=cfg.get("default_max_hops", 4),
            threshold_a=cfg.get("default_threshold_a", 50),
            threshold_b=cfg.get("default_threshold_b", 150),
            service_time_min=cfg.get("default_service_time_min", 120),
            min_vehicle_ft=20.0,   # Phase 2: enforce 20ft minimum
        )
    else:
        # Phase 2: enforce 20ft minimum — copy so original mh_configs is not mutated
        mh_cfg = dataclasses.replace(mh_cfg, min_vehicle_ft=20.0)

    dh_df = location_df[location_df["destination_hub_key"].isin(dh_keys)].copy()
    missing_loc = set(dh_keys) - set(dh_df["destination_hub_key"])
    if missing_loc:
        extras = location_df[location_df["destination_hub_key"].isin(missing_loc)].copy()
        dh_df  = pd.concat([dh_df, extras], ignore_index=True)

    if dh_df.empty:
        return _empty_mh_result(mh_name)

    dh_df = dh_df.copy()
    dh_df["current_fc_mh"] = mh_name

    return p4.run_agent4_for_mh(
        mh_name=mh_name,
        mh_cfg=mh_cfg,
        dh_df=dh_df,
        dist_dict=dist_dict,
        latlong=latlong,
        cfg=cfg,
        residual_threshold=residual_threshold,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Phase 5: Optimization — per-DH proxy cost (for LNS repair only)
# ─────────────────────────────────────────────────────────────────────────────

def _dh_direct_cost(
    dh:          str,
    mh_name:     str,
    dist_dict:   dict,
    mh_configs:  dict,
    location_df: pd.DataFrame,
    cfg:         dict,
) -> float:
    """
    Per-DH direct-routing cost, mirroring Agent 4's single-DH cost model exactly.
    Used ONLY for the LNS greedy-repair step (deciding which MH is cheaper for
    one destroyed DH in isolation).  All candidate-assignment evaluations use
    _run_a4_subset (full ILP) instead.

    Mirrors Agent 4 lines ~727-739:
      freq=1 : dist × rate(v1) × 30,  floor ₹90 000,  only if v1 ≤ ML cap
      freq=2 : dist × rate(v2) × 15,  floor ₹90 000,  only if freq_allowed==1 & v2 ≤ cap
      winner : min(c1, 1.1 × c2)
    """
    mh_cfg = mh_configs.get(mh_name)
    if mh_cfg is None:
        return float("inf")

    dist = dist_dict.get((mh_name, dh)) or dist_dict.get((dh, mh_name))
    if dist is None:
        return float("inf")

    row = location_df[location_df["destination_hub_key"] == dh]
    if row.empty:
        return float("inf")

    r            = row.iloc[0]
    demand       = float(r.get("total_cft",        0) or 0)
    ml_cap       = float(r.get("ML",              10) or 10)
    top266       = float(r.get("top266_shipments",  0) or 0)
    freq_allowed = p4.derive_freq_allowed(top266)

    threshold = cfg.get("local_zonal_distance_threshold_km", 200)
    rate_card = mh_cfg.local_rate_card if dist <= threshold else mh_cfg.zonal_rate_card

    _MIN_VEH = 20.0   # Phase 2: 20ft floor — consistent with _run_a4_subset
    v1 = max(p4.assign_vehicle_length(demand), _MIN_VEH)
    if v1 <= ml_cap:
        c1 = max(dist * rate_card.get(v1, 999.0) * 30, 90_000)
    else:
        c1 = float("inf")

    if freq_allowed == 1:
        v2 = max(p4.assign_vehicle_length(demand * 2), _MIN_VEH)
        if v2 <= ml_cap:
            c2 = max(dist * rate_card.get(v2, 999.0) * 15, 90_000)
        else:
            c2 = float("inf")
    else:
        c2 = float("inf")

    if c1 == float("inf") and c2 == float("inf"):
        return float("inf")

    return min(c1, 1.1 * c2)


# ─────────────────────────────────────────────────────────────────────────────
# Phase 5: Optimization — full Agent 4 cost for a candidate assignment
# ─────────────────────────────────────────────────────────────────────────────

def _a4_assignment_cost(
    assignment:         dict[str, str],
    pool_dhs:           list[str],
    mh1:                str,
    mh2:                str,
    dist_dict:          dict,
    latlong:            dict,
    mh_configs:         dict,
    location_df:        pd.DataFrame,
    cfg:                dict,
    residual_threshold: float,
    mhmh_monthly_mh1:   Optional[dict[str, float]] = None,
    mhmh_monthly_mh2:   Optional[dict[str, float]] = None,
) -> tuple[float, p4.Agent4MHResult, p4.Agent4MHResult]:
    """
    Run Agent 4's full routing pipeline for both MHs given this assignment.

    Scoring = MHDH routing cost (Agent 4 ILP) + MHMH trunk cost (pre-computed).
    When mhmh_monthly_mh1 / mhmh_monthly_mh2 are provided, each iteration
    minimises the true end-to-end cost so trunk-expensive moves are not selected
    purely on the basis of shorter last-mile routes.

    Returns (total_cost, result_mh1, result_mh2).
    total_cost includes MHMH for selection; r1/r2 carry MHDH-only results for
    reporting (MHDH and MHMH are kept separate in the output sheets).
    """
    from_dhs = [d for d in pool_dhs if assignment.get(d, mh1) == mh1]
    to_dhs   = [d for d in pool_dhs if assignment.get(d, mh1) == mh2]
    r1 = _run_a4_subset(mh1, from_dhs, dist_dict, latlong,
                        mh_configs, location_df, cfg, residual_threshold)
    r2 = _run_a4_subset(mh2, to_dhs,   dist_dict, latlong,
                        mh_configs, location_df, cfg, residual_threshold)
    mhdh_total = r1.total_monthly_cost + r2.total_monthly_cost

    # Add pre-computed MHMH trunk cost per DH at its assigned MH.
    # Missing values default to 0.0 (conservative — no phantom penalty/reward).
    mhmh_total = 0.0
    if mhmh_monthly_mh1 is not None or mhmh_monthly_mh2 is not None:
        _m1 = mhmh_monthly_mh1 or {}
        _m2 = mhmh_monthly_mh2 or {}
        for dh in pool_dhs:
            if assignment.get(dh, mh1) == mh1:
                mhmh_total += _m1.get(dh, 0.0)
            else:
                mhmh_total += _m2.get(dh, 0.0)

    return mhdh_total + mhmh_total, r1, r2


def optimize_pool_assignment(
    pool_dhs:           list[str],
    mh1:                str,
    mh2:                str,
    dist_dict:          dict,
    latlong:            dict,
    mh_configs:         dict,
    location_df:        pd.DataFrame,
    cfg:                dict,
    residual_threshold: float = 100.0,
    initial_assignment: Optional[dict[str, str]] = None,
    on_progress:        Optional[Callable[[str], None]] = None,
    mhmh_monthly_mh1:   Optional[dict[str, float]] = None,
    mhmh_monthly_mh2:   Optional[dict[str, float]] = None,
) -> tuple[dict[str, str], p4.Agent4MHResult, p4.Agent4MHResult]:
    """
    Find the cheapest assignment of pool_dhs between mh1 and mh2.

    Issue 1 / Issue 5 fix: ALL candidate evaluations (both enumeration and LNS)
    use _run_a4_subset (full Agent 4 ILP), not a simplified proxy.

    Scoring = MHDH routing cost + MHMH trunk cost (pre-computed per DH).
    mhmh_monthly_mh1 / mhmh_monthly_mh2 are {dh: monthly_cost_rs} dicts built
    before this call so every iteration sees true end-to-end cost.

    Strategy
    --------
    |pool| ≤ FULL_ENUM_THRESHOLD (15) : full enumeration, 2^N ILP evaluations
    |pool| > 15                        : LNS 500 iterations
      - Repair  : _dh_direct_cost + MHMH per destroyed DH (MHMH-aware greedy)
      - Evaluate: _a4_assignment_cost on the full repaired assignment (full ILP)

    Returns (best_assignment, best_result_mh1, best_result_mh2).
    The returned Agent 4 results are reused as the "after" outputs —
    no second Agent 4 run needed.
    """
    def _prog(msg: str) -> None:
        if on_progress:
            on_progress(msg)

    n = len(pool_dhs)
    if n == 0:
        return {}, _empty_mh_result(mh1), _empty_mh_result(mh2)

    base = {dh: (initial_assignment or {}).get(dh, mh1) for dh in pool_dhs}

    best_cost, best_r1, best_r2 = _a4_assignment_cost(
        base, pool_dhs, mh1, mh2,
        dist_dict, latlong, mh_configs, location_df, cfg, residual_threshold,
        mhmh_monthly_mh1=mhmh_monthly_mh1, mhmh_monthly_mh2=mhmh_monthly_mh2,
    )
    best = dict(base)

    if n <= FULL_ENUM_THRESHOLD:
        # ── Full enumeration — 2^N assignments, full ILP each ─────────────────
        total_combos = 1 << n
        _prog(f"  Enumeration: {total_combos} combinations to evaluate ...")
        for bits in range(total_combos):
            candidate = {
                pool_dhs[i]: (mh2 if (bits >> i) & 1 else mh1)
                for i in range(n)
            }
            cost, r1, r2 = _a4_assignment_cost(
                candidate, pool_dhs, mh1, mh2,
                dist_dict, latlong, mh_configs, location_df, cfg, residual_threshold,
                mhmh_monthly_mh1=mhmh_monthly_mh1, mhmh_monthly_mh2=mhmh_monthly_mh2,
            )
            if cost < best_cost:
                best_cost, best_r1, best_r2 = cost, r1, r2
                best = dict(candidate)
            # Progress every 10 combinations
            if (bits + 1) % 10 == 0 or (bits + 1) == total_combos:
                _prog(f"    Evaluating combination {bits + 1} of {total_combos}  "
                      f"(best so far: Rs {best_cost:,.0f})")
    else:
        # ── LNS: greedy repair + full-ILP evaluation ──────────────────────────
        rng = random.Random(42)
        k   = max(2, n // 5)   # destroy k DHs per iteration
        _prog(f"  LNS: {LNS_ITERATIONS} iterations, destroying {k} DHs per iteration ...")

        _lns_m1 = mhmh_monthly_mh1 or {}
        _lns_m2 = mhmh_monthly_mh2 or {}
        for lns_iter in range(LNS_ITERATIONS):
            destroyed = rng.sample(pool_dhs, k)
            candidate = dict(best)

            # Repair: assign each destroyed DH to the cheaper MH.
            # Proxy cost includes MHMH trunk so repair is also end-to-end aware.
            for dh in destroyed:
                c_mh1 = (_dh_direct_cost(dh, mh1, dist_dict, mh_configs, location_df, cfg)
                         + _lns_m1.get(dh, 0.0))
                c_mh2 = (_dh_direct_cost(dh, mh2, dist_dict, mh_configs, location_df, cfg)
                         + _lns_m2.get(dh, 0.0))
                candidate[dh] = mh1 if c_mh1 <= c_mh2 else mh2

            # Evaluate the repaired solution with full Agent 4 (Issue 5 fix)
            cost, r1, r2 = _a4_assignment_cost(
                candidate, pool_dhs, mh1, mh2,
                dist_dict, latlong, mh_configs, location_df, cfg, residual_threshold,
                mhmh_monthly_mh1=mhmh_monthly_mh1, mhmh_monthly_mh2=mhmh_monthly_mh2,
            )
            if cost < best_cost:
                best_cost, best_r1, best_r2 = cost, r1, r2
                best = dict(candidate)
            # Progress every 50 iterations
            if (lns_iter + 1) % 50 == 0:
                _prog(f"    LNS iteration {lns_iter + 1} of {LNS_ITERATIONS}, "
                      f"best cost so far: Rs {best_cost:,.0f}")

    return best, best_r1, best_r2


# ─────────────────────────────────────────────────────────────────────────────
# Issue 2: MH-MH delta resolver
# ─────────────────────────────────────────────────────────────────────────────

def _mhmh_cost_at_mh(
    dh:           str,
    target_mh:    str,
    agent3_df:    pd.DataFrame,
    mhmh_cache:   Optional[dict[tuple[str, str], Optional[float]]] = None,
) -> tuple[Optional[float], str]:
    """
    Resolve the MH-MH cost for dh if served by target_mh.

    Lookup chain
    ------------
    1. Agent 3 candidate columns (candidate_1 … candidate_4):
       exact candidate_X_mhmh_cost_rs written by Agent 3 Phase 1.
    2. Gap-fill cache: on-demand recompute via compute_mhmh_for_pairs()
       (for route-mates or DHs whose target MH was not in the top-4 candidates).
    3. Unavailable — returns (None, note).  No proxy approximations.

    Returns
    -------
    (cost_rs | None, note_string)
    """
    row = agent3_df[agent3_df["destination_hub_key"] == dh]
    if row.empty:
        return None, "DH not in Agent 3 output"

    r = row.iloc[0]

    # ── Step 1: exact MH-MH cost from Agent 3 candidate columns ─────────────
    for i in range(1, 5):
        cand_mh = str(r.get(f"candidate_{i}", "") or "").strip()
        if cand_mh.upper() == target_mh.strip().upper():
            mhmh_cost = pd.to_numeric(r.get(f"candidate_{i}_mhmh_cost_rs"), errors="coerce")
            if pd.notna(mhmh_cost):
                return float(mhmh_cost), f"Agent 3 candidate_{i}_mhmh_cost_rs"

    # ── Step 2: gap-fill cache (route-mates / non-candidate target MH) ───────
    if mhmh_cache is not None:
        cached = mhmh_cache.get((dh, target_mh))
        if cached is not None:
            return float(cached), "on-demand recompute (gap-fill)"

    # ── Step 3: unavailable ───────────────────────────────────────────────────
    return None, "unavailable"


# ─────────────────────────────────────────────────────────────────────────────
# Issue 3: MHDH Before — build H2H-grouped route rows
# ─────────────────────────────────────────────────────────────────────────────

def _build_mhdh_before_df(
    from_mh:     str,
    pool_dhs:    list[str],
    h2h_df:      pd.DataFrame,
    location_df: pd.DataFrame,
    dist_dict:   dict,
    mh_configs:  dict,
    cfg:         dict,
) -> pd.DataFrame:
    """
    Build MHDH Before sheet: one row per (Current_MH, MR_Number) group.

    Columns
    -------
    MR_Number | DHs_on_route | Current_MH | Distance_km | Vehicle_ft |
    Freq | Trips_per_month | Monthly_Cost_Rs | Route_Type

    For MR groups:    route = MH → DH1 → DH2 → … → MH (H2H order)
    For DIRECT DHs:   one row per DH with MR_Number = "DIRECT"
    For DHs with no H2H entry: one row per DH with MR_Number = "UNKNOWN"

    Vehicle sizing and cost use Agent 4's exact model (same as _dh_direct_cost).
    """
    mh_upper  = from_mh.strip().upper()
    mh_h2h    = h2h_df[h2h_df["Src_upper"] == mh_upper]
    pool_set  = set(pool_dhs)
    pool_upper = {d.strip().upper(): d for d in pool_dhs}   # upper → orig casing
    mh_cfg     = mh_configs.get(from_mh)
    threshold  = cfg.get("local_zonal_distance_threshold_km", 200)

    def _route_cost(route_dhs: list[str], dist: Optional[float]) -> tuple[Optional[int], Optional[int], Optional[float]]:
        """Return (freq, trips_per_month, monthly_cost_rs) or (None, None, None)."""
        if dist is None or mh_cfg is None:
            return None, None, None

        total_demand = sum(
            float(location_df[location_df["destination_hub_key"] == d].iloc[0].get("total_cft", 0) or 0)
            if not location_df[location_df["destination_hub_key"] == d].empty else 0.0
            for d in route_dhs
        )
        ml_cap = min(
            (float(location_df[location_df["destination_hub_key"] == d].iloc[0].get("ML", 10) or 10)
             if not location_df[location_df["destination_hub_key"] == d].empty else 10.0)
            for d in route_dhs
        ) if route_dhs else 10.0
        top266_total = sum(
            float(location_df[location_df["destination_hub_key"] == d].iloc[0].get("top266_shipments", 0) or 0)
            if not location_df[location_df["destination_hub_key"] == d].empty else 0.0
            for d in route_dhs
        )
        freq_allowed = p4.derive_freq_allowed(top266_total)
        rate_card    = mh_cfg.local_rate_card if dist <= threshold else mh_cfg.zonal_rate_card

        v1 = p4.assign_vehicle_length(total_demand)
        c1 = max(dist * rate_card.get(v1, 999.0) * 30, 90_000) if v1 <= ml_cap else float("inf")

        if freq_allowed == 1:
            v2 = p4.assign_vehicle_length(total_demand * 2)
            c2 = max(dist * rate_card.get(v2, 999.0) * 15, 90_000) if v2 <= ml_cap else float("inf")
        else:
            c2 = float("inf")

        if c1 == float("inf") and c2 == float("inf"):
            return None, None, None

        use_freq2 = (c2 != float("inf") and (1.1 * c2) < c1)
        freq      = 2 if use_freq2 else 1
        trips     = 15 if freq == 2 else 30
        cost      = round(1.1 * c2 if use_freq2 else c1, 2)
        return freq, trips, cost

    def _route_distance(nodes: list[str]) -> Optional[float]:
        """Sum of consecutive leg distances for nodes[0] → nodes[1] → … → nodes[-1]."""
        total = 0.0
        for i in range(len(nodes) - 1):
            d = dist_dict.get((nodes[i], nodes[i + 1])) or dist_dict.get((nodes[i + 1], nodes[i]))
            if d is None:
                return None
            total += d
        return round(total, 2)

    rows: list[dict] = []
    assigned_dhs: set[str] = set()   # track which pool DHs have been placed in a row

    # ── Non-DIRECT MR groups ──────────────────────────────────────────────────
    for mr_code, grp_df in mh_h2h.groupby("MR_cleaned"):
        if str(mr_code).startswith("DIRECT_"):
            continue

        # Pool DHs in this MR group (preserve H2H row order)
        ordered = []
        seen_in_grp: set[str] = set()
        for dest_upper in grp_df["Dest_upper"].tolist():
            if dest_upper in pool_upper and dest_upper not in seen_in_grp:
                ordered.append(pool_upper[dest_upper])
                seen_in_grp.add(dest_upper)
        if not ordered:
            continue

        # Route: MH → DH1 → … → DHn → MH
        nodes    = [from_mh] + ordered + [from_mh]
        dist_km  = _route_distance(nodes)
        freq, trips, cost = _route_cost(ordered, dist_km)

        total_demand = sum(
            float(location_df[location_df["destination_hub_key"] == d].iloc[0].get("total_cft", 0) or 0)
            if not location_df[location_df["destination_hub_key"] == d].empty else 0.0
            for d in ordered
        )
        veh_ft = p4.assign_vehicle_length(total_demand)

        rows.append({
            "MR_Number":      mr_code,
            "DHs_on_route":   ", ".join(ordered),
            "Current_MH":     from_mh,
            "Distance_km":    dist_km,
            "Vehicle_ft":     veh_ft,
            "Freq":           freq,
            "Trips_per_month": trips,
            "Monthly_Cost_Rs": cost,
            "Route_Type":     "Milkrun" if len(ordered) > 1 else "FTL_Dedicated",
        })
        assigned_dhs.update(ordered)

    # ── DIRECT and no-H2H DHs — one row per DH ───────────────────────────────
    for dh in pool_dhs:
        if dh in assigned_dhs:
            continue
        # Find MR code for this DH in H2H
        dh_upper = dh.strip().upper()
        h2h_row  = mh_h2h[mh_h2h["Dest_upper"] == dh_upper]
        mr_label = "UNKNOWN"
        if not h2h_row.empty:
            mr_label = str(h2h_row.iloc[0]["MR_cleaned"])
            if not mr_label.startswith("DIRECT_"):
                mr_label = mr_label   # non-DIRECT but not grouped (edge case)
            else:
                mr_label = "DIRECT"

        dist_km  = _route_distance([from_mh, dh, from_mh])
        freq, trips, cost = _route_cost([dh], dist_km)

        loc_row = location_df[location_df["destination_hub_key"] == dh]
        demand  = float(loc_row.iloc[0].get("total_cft", 0) or 0) if not loc_row.empty else 0.0
        veh_ft  = p4.assign_vehicle_length(demand)

        rows.append({
            "MR_Number":      mr_label,
            "DHs_on_route":   dh,
            "Current_MH":     from_mh,
            "Distance_km":    dist_km,
            "Vehicle_ft":     veh_ft,
            "Freq":           freq,
            "Trips_per_month": trips,
            "Monthly_Cost_Rs": cost,
            "Route_Type":     "FTL_Dedicated",
        })

    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────
# Phase 2 map — Leaflet HTML (mirrors Agent 3 style, pool DHs only)
# ─────────────────────────────────────────────────────────────────────────────

def _build_phase2_map_html(
    pr:               "MHPairResult",
    latlong:          dict[str, tuple[float, float]],
    location_df:      pd.DataFrame,
    per_dh_detail_df: Optional[pd.DataFrame] = None,
) -> str:
    """
    Generate a self-contained Leaflet HTML network map for one Phase 2 MH pair.

    Shows only pool DHs (flagged + route-mates).
    - Filled circle  : DH originally flagged by Agent 3
    - Ring circle    : route-mate added by H2H expansion
    - Blue  (from_mh): DH stays / before view
    - Orange (to_mh) : DH moves  / after view
    Before/After toggle in sidebar.
    """
    import colorsys as _cs
    import json as _json

    def _hsl(h: float, s: float = 0.65, l: float = 0.50) -> str:
        r, g, b = _cs.hls_to_rgb(h, l, s)
        return f"#{int(r*255):02x}{int(g*255):02x}{int(b*255):02x}"

    from_color = _hsl(0.60)   # blue
    to_color   = _hsl(0.07)   # orange

    ll_upper: dict[str, tuple[float, float]] = {
        k.strip().upper(): v for k, v in latlong.items()
    }

    def _pos(name: str) -> Optional[tuple[float, float]]:
        return ll_upper.get(name.strip().upper())

    # MH positions
    mh_data: dict[str, Any] = {}
    for mh, col in ((pr.from_mh, from_color), (pr.to_mh, to_color)):
        p = _pos(mh)
        if p:
            mh_data[mh] = {"lat": round(p[0], 6), "lon": round(p[1], 6), "color": col}

    # Build location lookup
    loc_idx: dict[str, pd.Series] = {}
    if "destination_hub_key" in location_df.columns:
        for _, lr in location_df.iterrows():
            loc_idx[str(lr["destination_hub_key"]).strip().upper()] = lr

    def _sf(v: Any) -> Optional[float]:
        try:
            return round(float(str(v).replace(",", "")), 2)
        except (ValueError, TypeError):
            return None

    flagged_upper = {d.strip().upper() for d in pr.flagged_dhs}
    moved_upper   = {d.strip().upper() for d in pr.dhs_moved}

    # Per-DH cost detail lookup (from Per_DH_Detail sheet)
    detail_idx: dict[str, dict] = {}
    if per_dh_detail_df is not None and not per_dh_detail_df.empty:
        for _, dr in per_dh_detail_df.iterrows():
            dk = str(dr.get("DH", "")).strip().upper()
            detail_idx[dk] = {
                "route":         dr.get("Route"),
                "mhdh_current":  _sf(dr.get("MHDH_Cost_Current")),
                "mhdh_assigned": _sf(dr.get("MHDH_Cost_Assigned")),
                "mhdh_saving":   _sf(dr.get("MHDH_Saving")),
                "mhmh_current":  _sf(dr.get("MHMH_Cost_Current")),
                "mhmh_assigned": _sf(dr.get("MHMH_Cost_Assigned")),
                "mhmh_saving":   _sf(dr.get("MHMH_Saving")),
                "total_saving":  _sf(dr.get("Total_Saving")),
            }

    dh_rows: list[dict] = []
    for dh in pr.pool_dhs:
        p = _pos(dh)
        if p is None:
            continue
        dh_u     = dh.strip().upper()
        is_flag  = dh_u in flagged_upper
        is_moved = dh_u in moved_upper
        loc      = loc_idx.get(dh_u, pd.Series(dtype=object))
        det      = detail_idx.get(dh_u, {})
        dh_rows.append({
            "name":          dh,
            "lat":           round(p[0], 6),
            "lon":           round(p[1], 6),
            "flagged":       is_flag,
            "moved":         is_moved,
            "after_mh":      pr.to_mh if is_moved else pr.from_mh,
            "cft":           _sf(loc.get("total_cft")),
            "ml":            _sf(loc.get("ML")),
            "top266":        _sf(loc.get("top266_shipments")),
            "route":         det.get("route"),
            "mhdh_current":  det.get("mhdh_current"),
            "mhdh_assigned": det.get("mhdh_assigned"),
            "mhdh_saving":   det.get("mhdh_saving"),
            "mhmh_current":  det.get("mhmh_current"),
            "mhmh_assigned": det.get("mhmh_assigned"),
            "mhmh_saving":   det.get("mhmh_saving"),
            "total_saving":  det.get("total_saving"),
        })

    n_dh      = len(dh_rows)
    n_moved   = sum(1 for d in dh_rows if d["moved"])
    n_stayed  = n_dh - n_moved
    n_mh      = len(mh_data)
    _mhmh_val = pr.mhmh_delta_rs if pr.mhmh_delta_rs is not None else 0.0
    _total    = pr.savings + _mhmh_val
    mhdh_fmt  = f"Rs {pr.savings:,.0f}"
    mhmh_fmt  = f"Rs {_mhmh_val:,.0f}" if pr.mhmh_delta_rs is not None else "N/A"
    total_fmt = f"Rs {_total:,.0f}"
    sav_fmt   = total_fmt   # kept for stats-bar compat
    sav_pct   = f"{100 * _total / pr.cost_before:.1f}" if pr.cost_before > 0 else "0.0"

    mh_data_js    = _json.dumps(mh_data,   ensure_ascii=False)
    dh_data_js    = _json.dumps(dh_rows,   ensure_ascii=False)
    from_mh_js    = _json.dumps(pr.from_mh)
    to_mh_js      = _json.dumps(pr.to_mh)
    from_color_js = _json.dumps(from_color)
    to_color_js   = _json.dumps(to_color)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Phase 2 Pool Map &mdash; {pr.from_mh} &rarr; {pr.to_mh}</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<style>
  *{{margin:0;padding:0;box-sizing:border-box;}}
  body{{font-family:"Segoe UI",sans-serif;background:#f0f2f5;color:#2d3748;display:flex;height:100vh;overflow:hidden;}}
  #sidebar{{width:290px;min-width:250px;max-width:320px;background:#fff;display:flex;flex-direction:column;border-right:1px solid #dde3ea;overflow:hidden;box-shadow:2px 0 8px rgba(0,0,0,.06);}}
  #sidebar-header{{padding:14px 16px;background:#f8fafc;border-bottom:1px solid #dde3ea;}}
  #sidebar-header h1{{font-size:14px;font-weight:700;color:#1a202c;margin-bottom:2px;}}
  #sidebar-header p{{font-size:11px;color:#718096;}}
  #view-toggle{{display:flex;gap:6px;padding:10px 12px 6px;}}
  .view-btn{{flex:1;padding:6px;font-size:11px;border-radius:6px;border:1px solid #dde3ea;cursor:pointer;background:#f8fafc;color:#718096;transition:all .15s;font-weight:500;}}
  .view-btn.active{{background:#3b82f6;color:#fff;border-color:#3b82f6;font-weight:700;}}
  .view-btn:hover:not(.active){{background:#edf2f7;color:#2d3748;}}
  #legend-box{{padding:10px 14px;border-bottom:1px solid #dde3ea;}}
  .leg-row{{display:flex;align-items:center;gap:8px;padding:3px 0;font-size:11px;color:#4a5568;}}
  .leg-dot{{width:12px;height:12px;border-radius:50%;flex-shrink:0;}}
  .leg-ring{{width:12px;height:12px;border-radius:50%;flex-shrink:0;border:2px solid #aaa;background:transparent;}}
  #stats-box{{padding:10px 14px;border-bottom:1px solid #dde3ea;}}
  .stat-row{{display:flex;justify-content:space-between;font-size:11px;padding:2px 0;}}
  .stat-row span:first-child{{color:#718096;}}
  .stat-row span:last-child{{color:#2d3748;font-weight:600;}}
  .stat-row.saving span:last-child{{color:#16a34a;}}
  #dh-list-box{{overflow-y:auto;flex:1;padding:0 8px 10px;}}
  .dh-list-header{{font-size:10px;font-weight:700;color:#a0aec0;text-transform:uppercase;letter-spacing:.05em;padding:8px 4px 4px;}}
  .dh-item{{display:flex;align-items:center;gap:7px;padding:4px 6px;border-radius:5px;font-size:11px;cursor:pointer;transition:background .1s;}}
  .dh-item:hover{{background:#f0f4f8;}}
  .dh-item-dot{{width:10px;height:10px;border-radius:50%;flex-shrink:0;}}
  .dh-item-ring{{width:10px;height:10px;border-radius:50%;flex-shrink:0;border:2px solid currentColor;background:transparent;}}
  .dh-name{{flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;color:#2d3748;font-size:10.5px;}}
  .dh-badge{{font-size:9px;padding:1px 6px;border-radius:8px;font-weight:700;flex-shrink:0;}}
  .badge-moved{{background:#fff3e0;color:#e65100;}}
  .badge-stayed{{background:#e8f5e9;color:#2e7d32;}}
  #map{{flex:1;}}
  #info-panel{{position:absolute;top:12px;right:12px;background:#fff;border:1px solid #dde3ea;border-radius:10px;padding:14px 16px;min-width:250px;max-width:310px;max-height:80vh;overflow-y:auto;display:none;z-index:1000;font-size:12px;box-shadow:0 4px 20px rgba(0,0,0,.12);}}
  #info-panel h3{{font-size:13px;font-weight:600;margin-bottom:8px;word-break:break-all;color:#1a202c;}}
  #info-panel .tag{{display:inline-block;padding:2px 9px;border-radius:12px;font-size:10px;font-weight:700;margin-bottom:7px;color:#fff;}}
  .info-table{{width:100%;border-collapse:collapse;margin-top:4px;}}
  .info-table tr{{border-bottom:1px solid #edf2f7;}}
  .info-table tr:last-child{{border-bottom:none;}}
  .info-table td{{padding:6px 4px;font-size:11px;}}
  .info-table td:first-child{{color:#718096;width:55%;}}
  .info-table td:last-child{{color:#2d3748;font-weight:600;text-align:right;}}
  .info-mh-row{{display:flex;align-items:center;gap:6px;padding:6px 4px;border-bottom:1px solid #edf2f7;font-size:11px;margin-bottom:4px;}}
  .info-mh-dot{{width:10px;height:10px;border-radius:50%;flex-shrink:0;}}
  .info-mh-name{{color:#2d3748;font-weight:600;word-break:break-all;}}
  #info-close{{position:absolute;top:8px;right:10px;cursor:pointer;color:#a0aec0;font-size:18px;line-height:1;}}
  #info-close:hover{{color:#2d3748;}}
  #stats-bar{{position:absolute;bottom:20px;right:12px;background:#fff;border:1px solid #dde3ea;border-radius:8px;padding:7px 14px;z-index:1000;font-size:11px;color:#718096;display:flex;gap:16px;box-shadow:0 2px 8px rgba(0,0,0,.08);}}
  #stats-bar span{{color:#2d3748;font-weight:600;}}
  .leaflet-tooltip{{background:#fff;border:1px solid #dde3ea;color:#2d3748;font-size:11px;padding:4px 8px;border-radius:5px;box-shadow:0 2px 6px rgba(0,0,0,.1);}}
  .leaflet-tooltip::before{{display:none;}}
</style>
</head>
<body>
<div id="sidebar">
  <div id="sidebar-header">
    <h1 id="pair-title"></h1>
    <p>Phase 2 pool — {n_dh} DHs ({n_moved} moved, {n_stayed} stayed)</p>
  </div>

  <div id="view-toggle">
    <button class="view-btn" id="btn-before" onclick="setView('before')">Before</button>
    <button class="view-btn active" id="btn-after" onclick="setView('after')">After (recommended)</button>
  </div>

  <div id="legend-box">
    <div style="font-size:10px;color:#a0aec0;font-weight:700;text-transform:uppercase;letter-spacing:.05em;margin-bottom:5px">Legend</div>
    <div class="leg-row"><div class="leg-dot" id="leg-from"></div><span id="leg-from-label"></span></div>
    <div class="leg-row"><div class="leg-dot" id="leg-to"></div><span id="leg-to-label"></span></div>
    <div class="leg-row" style="margin-top:4px;border-top:1px solid #edf2f7;padding-top:4px">
      <div class="leg-dot" style="background:#666"></div><span style="font-size:10.5px">Flagged by Agent 3 (filled)</span>
    </div>
    <div class="leg-row">
      <div class="leg-ring"></div><span style="font-size:10.5px">Route-mate via H2H (ring)</span>
    </div>
  </div>

  <div id="stats-box">
    <div class="stat-row"><span>Pool DHs</span><span>{n_dh}</span></div>
    <div class="stat-row"><span>Moved to {pr.to_mh}</span><span>{n_moved}</span></div>
    <div class="stat-row"><span>Stayed at {pr.from_mh}</span><span>{n_stayed}</span></div>
    <div class="stat-row"><span>MHDH Saving / month</span><span>{mhdh_fmt}</span></div>
    <div class="stat-row"><span>MHMH Saving / month</span><span>{mhmh_fmt}</span></div>
    <div class="stat-row saving"><span>Total Saving / month</span><span>{total_fmt} ({sav_pct}%)</span></div>
  </div>

  <div id="dh-list-box">
    <div class="dh-list-header">Pool DHs</div>
    <div id="dh-list"></div>
  </div>
</div>

<div id="map"></div>

<div id="info-panel">
  <span id="info-close">&times;</span>
  <div id="info-tag" class="tag"></div>
  <h3 id="info-name"></h3>
  <div id="info-body"></div>
</div>

<div id="stats-bar">
  <div>MHs: <span>{n_mh}</span></div>
  <div>Pool: <span>{n_dh}</span></div>
  <div>Moved: <span id="stat-moved">{n_moved}</span></div>
  <div>Stayed: <span id="stat-stayed">{n_stayed}</span></div>
</div>

<script>
const MH_DATA   = {mh_data_js};
const DH_DATA   = {dh_data_js};
const FROM_MH   = {from_mh_js};
const TO_MH     = {to_mh_js};
const FROM_COLOR = {from_color_js};
const TO_COLOR   = {to_color_js};

let currentView = 'after';
const dhMarkerRefs = [];
const mhMarkerRefs = [];

// Pair title
document.getElementById('pair-title').textContent = FROM_MH + ' → ' + TO_MH;
document.getElementById('leg-from').style.background = FROM_COLOR;
document.getElementById('leg-from-label').textContent = FROM_MH + ' (stays)';
document.getElementById('leg-to').style.background = TO_COLOR;
document.getElementById('leg-to-label').textContent = TO_MH + ' (moved to)';

const map = L.map('map', {{center:[22.5,82.5],zoom:5,zoomControl:true,preferCanvas:true}});
L.tileLayer('https://{{s}}.basemaps.cartocdn.com/rastertiles/voyager/{{z}}/{{x}}/{{y}}{{r}}.png',{{
  attribution:'&copy; OpenStreetMap &copy; CARTO', maxZoom:18
}}).addTo(map);

function makeDHFilled(color, size) {{
  const s = size||10;
  const svg = `<svg xmlns='http://www.w3.org/2000/svg' width='${{s}}' height='${{s}}' viewBox='0 0 ${{s}} ${{s}}'><circle cx='${{s/2}}' cy='${{s/2}}' r='${{s/2-1}}' fill='${{color}}' stroke='white' stroke-width='1.5'/></svg>`;
  return L.divIcon({{html:svg,className:'',iconSize:[s,s],iconAnchor:[s/2,s/2]}});
}}
function makeDHRing(color, size) {{
  const s = size||10;
  const svg = `<svg xmlns='http://www.w3.org/2000/svg' width='${{s}}' height='${{s}}' viewBox='0 0 ${{s}} ${{s}}'><circle cx='${{s/2}}' cy='${{s/2}}' r='${{s/2-1}}' fill='white' stroke='${{color}}' stroke-width='2'/></svg>`;
  return L.divIcon({{html:svg,className:'',iconSize:[s,s],iconAnchor:[s/2,s/2]}});
}}
function makeMHIcon(color) {{
  const svg = `<svg xmlns='http://www.w3.org/2000/svg' width='22' height='28' viewBox='0 0 28 36'><path d='M14 0 C6.268 0 0 6.268 0 14 C0 24.5 14 36 14 36 C14 36 28 24.5 28 14 C28 6.268 21.732 0 14 0 Z' fill='${{color}}' stroke='white' stroke-width='1.5'/><path d='M14 6 L8 12 L9.5 12 L9.5 20 L12.5 20 L12.5 16 L15.5 16 L15.5 20 L18.5 20 L18.5 12 L20 12 Z' fill='white' opacity='0.9'/></svg>`;
  return L.divIcon({{html:svg,className:'',iconSize:[22,28],iconAnchor:[11,28]}});
}}

function fmtNum(v) {{
  if (v===null||v===undefined) return '—';
  return typeof v==='number'?v.toLocaleString(undefined,{{maximumFractionDigits:1}}):v;
}}
function fmtRs(v) {{
  if (v===null||v===undefined) return '—';
  const n=Math.round(Number(v));
  if (isNaN(n)) return '—';
  return 'Rs '+n.toLocaleString();
}}
function savCell(v) {{
  if (v===null||v===undefined) return '<td style="text-align:right;font-weight:600">—</td>';
  const n=Number(v);
  const col=n>=0?'#16a34a':'#dc2626';
  return `<td style="text-align:right;font-weight:700;color:${{col}}">${{fmtRs(v)}}</td>`;
}}

function showInfo(name,isM,color,bodyHTML) {{
  const p=document.getElementById('info-panel');
  const t=document.getElementById('info-tag');
  t.textContent=isM?'MOTHER HUB':'DESTINATION HUB';
  t.style.background=color;
  document.getElementById('info-name').textContent=name;
  document.getElementById('info-body').innerHTML=bodyHTML;
  p.style.display='block';
}}
document.getElementById('info-close').onclick=()=>document.getElementById('info-panel').style.display='none';
map.on('click',()=>document.getElementById('info-panel').style.display='none');

// Add MH markers
Object.entries(MH_DATA).forEach(([mh,d])=>{{
  if(!d.lat||!d.lon) return;
  const m=L.marker([d.lat,d.lon],{{icon:makeMHIcon(d.color),zIndexOffset:1000}}).addTo(map);
  mhMarkerRefs.push({{mh,marker:m}});
  m.bindTooltip(`<b>${{mh}}</b>`,{{permanent:false,direction:'top'}});
  m.on('click',e=>{{
    L.DomEvent.stopPropagation(e);
    const before_dhs = DH_DATA.map(d=>d.name);
    const after_dhs  = DH_DATA.filter(d=>d.after_mh===mh).map(d=>d.name);
    const body=`<table class="info-table">
      <tr><td>DHs before</td><td>${{before_dhs.length}}</td></tr>
      <tr><td>DHs after</td><td>${{after_dhs.length}}</td></tr>
    </table>`;
    showInfo(mh,true,d.color,body);
  }});
}});

// Add DH markers + lines
const allLatLngs = [];
DH_DATA.forEach(dh=>{{
  if(!dh.lat||dh.lon===undefined) return;
  allLatLngs.push([dh.lat,dh.lon]);

  const afterColor  = dh.moved ? TO_COLOR : FROM_COLOR;
  const beforeColor = FROM_COLOR;
  const afterMHPos  = MH_DATA[dh.after_mh];
  const fromMHPos   = MH_DATA[FROM_MH];

  // lines: before (always to from_mh) and after (to assigned mh)
  let polyBefore=null, polyAfter=null;
  if(fromMHPos) polyBefore = L.polyline([[dh.lat,dh.lon],[fromMHPos.lat,fromMHPos.lon]],{{color:FROM_COLOR,weight:1.2,opacity:0.4}});
  if(afterMHPos) polyAfter = L.polyline([[dh.lat,dh.lon],[afterMHPos.lat,afterMHPos.lon]],{{color:afterColor,weight:1.2,opacity:0.4}});

  const markerBefore = L.marker([dh.lat,dh.lon],{{icon: dh.flagged ? makeDHFilled(FROM_COLOR,10) : makeDHRing(FROM_COLOR,10)}});
  const markerAfter  = L.marker([dh.lat,dh.lon],{{icon: dh.flagged ? makeDHFilled(afterColor,10) : makeDHRing(afterColor,10)}});

  dhMarkerRefs.push({{dh, markerBefore, markerAfter, polyBefore, polyAfter}});

  function _onClick(e, marker) {{
    L.DomEvent.stopPropagation(e);
    const c = currentView==='after' ? (dh.moved?TO_COLOR:FROM_COLOR) : FROM_COLOR;
    const routeRow = dh.route
      ? `<tr><td style="color:#718096;vertical-align:top">Route</td><td style="font-size:10px;word-break:break-all;max-width:160px;text-align:right">${{dh.route}}</td></tr>`
      : '';
    const body = `
      <div class="info-mh-row"><div class="info-mh-dot" style="background:${{c}}"></div><span class="info-mh-name">${{currentView==='after'?dh.after_mh:FROM_MH}}</span></div>
      <table class="info-table">
        <tr><td>Type</td><td style="text-align:right">${{dh.flagged?'Flagged (Agent 3)':'Route-mate (H2H)'}}</td></tr>
        <tr><td>Status</td><td style="text-align:right;color:${{dh.moved?'#e65100':'#2e7d32'}};font-weight:700">${{dh.moved?'Moved → '+TO_MH:'Stayed'}}</td></tr>
        <tr><td>Total CFT</td><td style="text-align:right">${{fmtNum(dh.cft)}}</td></tr>
        ${{routeRow}}
        <tr><td colspan="2" style="padding-top:8px;padding-bottom:2px;font-size:10px;font-weight:700;color:#718096;text-transform:uppercase;letter-spacing:.05em">MH-DH Cost</td></tr>
        <tr><td>Current</td><td style="text-align:right">${{fmtRs(dh.mhdh_current)}}</td></tr>
        <tr><td>Assigned</td><td style="text-align:right">${{fmtRs(dh.mhdh_assigned)}}</td></tr>
        <tr><td>Saving</td>${{savCell(dh.mhdh_saving)}}</tr>
        <tr><td colspan="2" style="padding-top:8px;padding-bottom:2px;font-size:10px;font-weight:700;color:#718096;text-transform:uppercase;letter-spacing:.05em">MH-MH Cost</td></tr>
        <tr><td>Current</td><td style="text-align:right">${{fmtRs(dh.mhmh_current)}}</td></tr>
        <tr><td>Assigned</td><td style="text-align:right">${{fmtRs(dh.mhmh_assigned)}}</td></tr>
        <tr><td>Saving</td>${{savCell(dh.mhmh_saving)}}</tr>
        <tr><td colspan="2" style="border-top:1px solid #e2e8f0;padding-top:4px"></td></tr>
        <tr><td><b>Total Saving</b></td>${{savCell(dh.total_saving)}}</tr>
      </table>`;
    showInfo(dh.name,false,c,body);
  }}
  markerBefore.on('click',e=>_onClick(e,markerBefore));
  markerAfter.on('click',e=>_onClick(e,markerAfter));
  markerBefore.on('mouseover',()=>markerBefore.bindTooltip(dh.name,{{permanent:false,direction:'top'}}).openTooltip());
  markerAfter.on('mouseover',()=>markerAfter.bindTooltip(dh.name,{{permanent:false,direction:'top'}}).openTooltip());
}});

function setView(v) {{
  currentView = v;
  document.getElementById('btn-before').classList.toggle('active', v==='before');
  document.getElementById('btn-after').classList.toggle('active',  v==='after');
  dhMarkerRefs.forEach(ref=>{{
    if(v==='after') {{
      if(ref.markerBefore._map) ref.markerBefore.remove();
      if(ref.polyBefore && ref.polyBefore._map) ref.polyBefore.remove();
      ref.markerAfter.addTo(map);
      if(ref.polyAfter) ref.polyAfter.addTo(map);
    }} else {{
      if(ref.markerAfter._map) ref.markerAfter.remove();
      if(ref.polyAfter && ref.polyAfter._map) ref.polyAfter.remove();
      ref.markerBefore.addTo(map);
      if(ref.polyBefore) ref.polyBefore.addTo(map);
    }}
  }});
}}

// Build DH sidebar list
const listEl = document.getElementById('dh-list');
DH_DATA.slice().sort((a,b)=>{{
  if(a.moved!==b.moved) return a.moved?-1:1;
  return (a.name||'').localeCompare(b.name||'');
}}).forEach(dh=>{{
  const color = dh.moved ? TO_COLOR : FROM_COLOR;
  const item = document.createElement('div');
  item.className='dh-item';
  item.innerHTML=`
    ${{dh.flagged
      ? `<div class="dh-item-dot" style="background:${{color}}"></div>`
      : `<div class="dh-item-ring" style="color:${{color}};border-color:${{color}}"></div>`}}
    <span class="dh-name" title="${{dh.name}}">${{dh.name}}</span>
    <span class="dh-badge ${{dh.moved?'badge-moved':'badge-stayed'}}">${{dh.moved?'moved':'stayed'}}</span>`;
  item.onclick=()=>{{
    const ref=dhMarkerRefs.find(r=>r.dh.name===dh.name);
    if(ref) {{
      const m=currentView==='after'?ref.markerAfter:ref.markerBefore;
      map.setView([dh.lat,dh.lon],10);
      m.fire('click');
    }}
  }};
  listEl.appendChild(item);
}});

// Initial: show after view
setView('after');

// Fit map to all DH + MH positions
const allPts = allLatLngs.concat(Object.values(MH_DATA).filter(d=>d.lat).map(d=>[d.lat,d.lon]));
if(allPts.length) map.fitBounds(L.latLngBounds(allPts).pad(0.12));
</script>
</body>
</html>"""


# ─────────────────────────────────────────────────────────────────────────────
# Phase 6: Build per-pair 5-sheet output
# ─────────────────────────────────────────────────────────────────────────────

def build_pair_output(
    from_mh:             str,
    to_mh:               str,
    pool_dhs:            list[str],
    best_assignment:     dict[str, str],
    agent3_df:           pd.DataFrame,
    location_df:         pd.DataFrame,
    h2h_df:              pd.DataFrame,
    dist_dict:           dict,
    mh_configs:          dict,
    cfg:                 dict,
    before_from_result:  p4.Agent4MHResult,
    before_to_result:    p4.Agent4MHResult,
    after_from_result:   p4.Agent4MHResult,
    after_to_result:     p4.Agent4MHResult,
    mhmh_cache:          Optional[dict[tuple[str, str], Optional[float]]] = None,
    flagged_dhs:         Optional[list[str]] = None,
) -> tuple[dict[str, pd.DataFrame], Optional[float], Optional[float]]:
    """
    Build the 5-sheet Excel content for one MH pair.

    Returns (sheets_dict, pool_mhmh_delta_monthly_rs, agent3_est_saving_rs)
      pool_mhmh_delta_monthly_rs — sum of (mhmh_current − mhmh_target) × 30 across moved
                                    DHs where both sides are available; None if all missing.
      agent3_est_saving_rs       — sum of cost_delta_rs × 30 for moved DHs (Agent 3 estimate).

    Sheets
    ------
    Summary          : MHDH delta + MHMH delta + Total_Savings (I13)
    MHDH_Before      : Agent 4 baseline ILP routes — exact costs
    MHDH_After       : Agent 4 final route assignments after optimisation
    MHMH_Cost_Impact : per-DH MH-MH cost breakdown (daily + monthly)
    Per_DH_Detail    : route-based MHDH costs (route_cost/n_dhs), MHMH both sides ×30,
                       Total_Saving per DH — sums match Summary (I13)
    """
    DAYS = DAYS_PER_MONTH

    pool_cost_before = (before_from_result.total_monthly_cost
                        + before_to_result.total_monthly_cost)
    pool_cost_after  = (after_from_result.total_monthly_cost
                        + after_to_result.total_monthly_cost)
    mhdh_delta = pool_cost_before - pool_cost_after

    dhs_moved  = [d for d in pool_dhs if best_assignment.get(d) == to_mh]
    dhs_stayed = [d for d in pool_dhs if best_assignment.get(d, from_mh) == from_mh]

    a3_pool = agent3_df[agent3_df["destination_hub_key"].isin(pool_dhs)]

    # ── DH → route mapping from ILP results ──────────────────────────────────
    # Each DH is in exactly one route in before/after results.
    # MHDH cost per DH = route monthly_cost / number of DHs on that route.
    def _build_route_map(results: list) -> dict:
        """Returns {dh: {route_sequence, monthly_cost (accumulated share), mh}}
        A DH may appear in multiple routes (e.g. milkrun + FTL when demand exceeds
        single vehicle capacity). Cost shares are accumulated, not overwritten."""
        m: dict = {}
        for res in results:
            df = res.final_assignment_df
            if df.empty:
                continue
            for _, row in df.iterrows():
                hubs = row.get("hubs", [])
                if not isinstance(hubs, list):
                    try:
                        import ast
                        hubs = ast.literal_eval(str(hubs))
                    except Exception:
                        hubs = [str(hubs)]
                n     = max(len(hubs), 1)
                share = float(row.get("monthly_cost", 0.0)) / n
                mh    = str(row.get("MH", ""))
                rseq  = str(row.get("route_sequence", ""))
                for dh in hubs:
                    dh_str = str(dh)
                    if dh_str in m:
                        m[dh_str]["monthly_cost"] += share
                        # append route sequence for multi-route DHs
                        if rseq and rseq not in m[dh_str]["route_sequence"]:
                            m[dh_str]["route_sequence"] += f" + {rseq}"
                    else:
                        m[dh_str] = {
                            "route_sequence": rseq,
                            "monthly_cost":   share,
                            "mh":             mh,
                        }
        return m

    before_route_map = _build_route_map([before_from_result, before_to_result])
    after_route_map  = _build_route_map([after_from_result,  after_to_result])

    # ── MHMH costs for all pool DHs — both current and target sides ──────────
    # Current side: scan candidate columns for from_mh (I13 fix — no longer
    # restricted to assigned_fc_mh == from_mh).
    # Target side:  existing _mhmh_cost_at_mh logic (candidate cols + gap-fill).
    # Both returned as per-day values; multiply by DAYS for monthly.
    mhmh_current_day: dict[str, Optional[float]] = {}
    mhmh_target_day:  dict[str, Optional[float]] = {}
    for dh in pool_dhs:
        mc, _ = _mhmh_cost_at_mh(dh, from_mh, agent3_df, mhmh_cache)
        mt, _ = _mhmh_cost_at_mh(dh, to_mh,   agent3_df, mhmh_cache)
        mhmh_current_day[dh] = mc
        mhmh_target_day[dh]  = mt

    # Per-DH MHMH monthly delta (positive = saving, i.e. current > target)
    # Stayed DHs: delta = 0 (their MH assignment doesn't change → no MH-MH cost change)
    mhmh_dh_delta: dict[str, Optional[float]] = {}
    for dh in pool_dhs:
        new_mh_check = best_assignment.get(dh, from_mh)
        if new_mh_check == from_mh:
            mhmh_dh_delta[dh] = 0.0   # stayed — no MHMH change
        else:
            mc, mt = mhmh_current_day[dh], mhmh_target_day[dh]
            mhmh_dh_delta[dh] = (mc - mt) * DAYS if (mc is not None and mt is not None) else None

    # Pool-level MHMH delta: sum of moved DHs where both sides are available
    avail = [v for v in mhmh_dh_delta.values() if v is not None]
    pool_mhmh_delta: Optional[float] = round(sum(avail), 2) if avail else None

    total_savings = mhdh_delta + (pool_mhmh_delta or 0.0)

    # Agent 3 estimated monthly saving: sum of cost_delta_rs × 30 for MOVED DHs only
    a3_est_vals = []
    for dh in dhs_moved:
        row_a3 = a3_pool[a3_pool["destination_hub_key"] == dh]
        if not row_a3.empty:
            v = pd.to_numeric(row_a3.iloc[0].get("cost_delta_rs"), errors="coerce")
            if pd.notna(v):
                a3_est_vals.append(float(v) * DAYS)
    agent3_est_saving_rs: Optional[float] = round(sum(a3_est_vals), 2) if a3_est_vals else None

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    summary_df = pd.DataFrame([{
        "From_MH":              from_mh,
        "To_MH":                to_mh,
        "Agent3_Flagged_DHs":   len(flagged_dhs) if flagged_dhs is not None else None,
        "Pool_DHs":             len(pool_dhs),
        "DHs_Moved":            len(dhs_moved),
        "MHDH_Saving_Rs":       round(mhdh_delta,   2),
        "MHMH_Saving_Rs":       pool_mhmh_delta,       # None if all unavailable
        "Total_Savings_Rs":     round(total_savings, 2),
        "Agent3_Est_Saving_Rs": agent3_est_saving_rs,
    }])

    # ── Sheet 2 & 3 helper ───────────────────────────────────────────────────
    def _tag_mh(result: p4.Agent4MHResult, mh: str) -> pd.DataFrame:
        df = result.final_assignment_df.copy()
        if not df.empty and "MH" not in df.columns:
            df.insert(0, "MH", mh)
        return df

    # ── Sheet 2: MHDH Before ─────────────────────────────────────────────────
    mhdh_before_df = pd.concat(
        [f for f in [_tag_mh(before_from_result, from_mh),
                     _tag_mh(before_to_result,   to_mh)] if not f.empty],
        ignore_index=True,
    )

    # ── Sheet 3: MHDH After ──────────────────────────────────────────────────
    mhdh_after_df = pd.concat(
        [f for f in [_tag_mh(after_from_result, from_mh),
                     _tag_mh(after_to_result,   to_mh)] if not f.empty],
        ignore_index=True,
    )

    # ── Sheet 4: MHMH Cost Impact ─────────────────────────────────────────────
    mhmh_rows: list[dict] = []
    for dh in pool_dhs:
        row    = a3_pool[a3_pool["destination_hub_key"] == dh]
        r      = row.iloc[0] if not row.empty else pd.Series(dtype=object)
        new_mh = best_assignment.get(dh, from_mh)
        mc_day = mhmh_current_day[dh]
        mt_day = mhmh_target_day[dh]
        delta_day = (mc_day - mt_day) if (mc_day is not None and mt_day is not None) else None
        mhmh_rows.append({
            "DH":                           dh,
            "Original_MH":                  from_mh,
            "Assigned_MH":                  new_mh,
            "Moved":                        new_mh != from_mh,
            "MH_MH_Cost_Current_Rs_Day":    mc_day,
            "MH_MH_Cost_Target_Rs_Day":     mt_day,
            "MH_MH_Delta_Rs_Day":           delta_day,
            "MH_MH_Delta_Rs_Month":         round(delta_day * DAYS, 2) if delta_day is not None else None,
        })
    mhmh_df = pd.DataFrame(mhmh_rows)

    # ── Sheet 5: Per DH Detail (I13) ─────────────────────────────────────────
    # MHDH cost = route_cost / n_dhs_on_route (exact ILP cost, not a proxy).
    # MHMH cost = daily × DAYS for monthly comparison.
    # Total_Saving = MHDH_Saving + MHMH_Saving (where available).
    # Sum of Total_Saving across all pool DHs == Total_Savings_Rs in Summary.
    detail_rows: list[dict] = []
    for dh in pool_dhs:
        r   = (a3_pool[a3_pool["destination_hub_key"] == dh].iloc[0]
               if not a3_pool[a3_pool["destination_hub_key"] == dh].empty
               else pd.Series(dtype=object))
        loc = (location_df[location_df["destination_hub_key"] == dh].iloc[0]
               if not location_df[location_df["destination_hub_key"] == dh].empty
               else pd.Series(dtype=object))
        new_mh = best_assignment.get(dh, from_mh)

        # MHDH: monthly_cost is already the per-DH cost share (accumulated in _build_route_map)
        b_info = before_route_map.get(dh, {})
        a_info = after_route_map.get(dh, {})
        mhdh_current  = b_info["monthly_cost"] if b_info else None
        mhdh_assigned = a_info["monthly_cost"] if a_info else None
        mhdh_saving   = ((mhdh_current - mhdh_assigned)
                         if mhdh_current is not None and mhdh_assigned is not None
                         else None)
        route_str = b_info.get("route_sequence", "")

        # MHMH: daily × DAYS
        mc_day = mhmh_current_day[dh]
        mt_day = mhmh_target_day[dh]
        mhmh_cost_current  = mc_day * DAYS if mc_day is not None else None
        mhmh_cost_assigned = mt_day * DAYS if mt_day is not None else None
        mhmh_saving        = mhmh_dh_delta[dh]   # already (current - target) × DAYS

        # Total saving = MHDH + MHMH (use 0 for missing MHMH so MHDH still shows)
        total_dh_saving = (mhdh_saving + (mhmh_saving or 0.0)
                           if mhdh_saving is not None else None)
        dh_saving_confirmed = (
            "Yes" if total_dh_saving is not None
                     and total_dh_saving > SAVING_CONFIRMED_THRESHOLD_RS
            else "No"
        )

        a3_delta_rs = pd.to_numeric(r.get("cost_delta_rs"), errors="coerce")
        detail_rows.append({
            "DH":                           dh,
            "Original_MH":                  from_mh,
            "Assigned_MH":                  new_mh,
            "Moved":                        new_mh != from_mh,
            "Total_CFT":                    pd.to_numeric(loc.get("total_cft"),        errors="coerce"),
            "ML":                           pd.to_numeric(loc.get("ML"),               errors="coerce"),
            "Top266_Shipments":             pd.to_numeric(loc.get("top266_shipments"), errors="coerce"),
            "Agent3_Est_Saving_Rs_Month":   round(float(a3_delta_rs) * DAYS, 2)
                                            if pd.notna(a3_delta_rs) else None,
            "Agent3_MH_Recommendation":     r.get("assigned_fc_mh"),
            "Route":                        route_str,
            "MHDH_Cost_Current":            round(mhdh_current,  2) if mhdh_current  is not None else None,
            "MHDH_Cost_Assigned":           round(mhdh_assigned, 2) if mhdh_assigned is not None else None,
            "MHDH_Saving":                  round(mhdh_saving,   2) if mhdh_saving   is not None else None,
            "MHMH_Cost_Current":            round(mhmh_cost_current,  2) if mhmh_cost_current  is not None else None,
            "MHMH_Cost_Assigned":           round(mhmh_cost_assigned, 2) if mhmh_cost_assigned is not None else None,
            "MHMH_Saving":                  round(mhmh_saving,        2) if mhmh_saving        is not None else None,
            "Total_Saving":                 round(total_dh_saving, 2)    if total_dh_saving     is not None else None,
            "Saving_confirmed":             dh_saving_confirmed,
        })
    detail_df = pd.DataFrame(detail_rows)

    return (
        {
            "Summary":          summary_df,
            "MHDH_Before":      mhdh_before_df,
            "MHDH_After":       mhdh_after_df,
            "MHMH_Cost_Impact": mhmh_df,
            "Per_DH_Detail":    detail_df,
        },
        pool_mhmh_delta,
        agent3_est_saving_rs,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Phase 7a: Formatted "Routes" sheet (side-by-side before/after)
# ─────────────────────────────────────────────────────────────────────────────

def _write_routes_sheet(wb, pr: "MHPairResult") -> None:
    """
    Add a formatted side-by-side Routes sheet to openpyxl workbook wb.

    Layout
    ------
    Row 1  : section headers — CURRENT ROUTES | NEW ROUTES
    Row 2  : column headers
    Row 3+ : route data rows (before left, after right)
    Total  : totals row
    Saving : MH-DH saving calculation
    """
    from openpyxl import Workbook as _WB          # noqa – already a dep
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Routes")

    # ── Style helpers ────────────────────────────────────────────────────────
    def _f(bold=False, sz=10, color="000000", italic=False):
        return Font(name="Arial", bold=bold, size=sz, color=color, italic=italic)

    def _fill(hex_col):
        return PatternFill("solid", fgColor=hex_col)

    def _al(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    _thin   = Side(style="thin", color="CCCCCC")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    T_FILL  = _fill("1F4E79")   # dark blue  — title
    B_FILL  = _fill("FCE4D6")   # light orange — before section header
    A_FILL  = _fill("E2EFDA")   # light green  — after section header
    H_FILL  = _fill("D6DCE4")   # grey-blue    — column headers
    TO_FILL = _fill("FFF2CC")   # yellow       — totals
    SV_FILL = _fill("C6EFCE")   # green        — saving
    W_FILL  = _fill("FFFFFF")   # white
    AL_FILL = _fill("F5F5F5")   # alternating row

    # ── Column widths  A-I = before (9 cols), J = spacer, K-S = after ────────
    widths = [58, 32, 24, 9, 10, 9, 7, 9, 18,
              3,
              58, 32, 24, 9, 10, 9, 7, 9, 18]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Extract route rows from Agent4MHResult objects ───────────────────────
    def _route_rows(results):
        rows = []
        for res in (results or []):
            if res is None:
                continue
            df = res.final_assignment_df
            if df is None or df.empty:
                continue
            for _, row in df.iterrows():
                hubs = row.get("hubs", [])
                if not isinstance(hubs, list):
                    try:
                        import ast as _ast
                        hubs = _ast.literal_eval(str(hubs))
                    except Exception:
                        hubs = [str(hubs)]
                freq   = float(row.get("Freq", 1) or 1)
                dist   = float(row.get("dist", 0) or 0)
                cost   = float(row.get("monthly_cost", 0) or 0)
                denom  = dist * freq * DAYS_PER_MONTH
                rspkm  = round(cost / denom, 1) if denom > 0 else None
                rtype  = str(row.get("Route_Type", "Milkrun")).lower()
                rows.append({
                    "route":  f"{row.get('route_sequence', '')}  ({rtype})",
                    "dhs":    ", ".join(str(h) for h in hubs),
                    "mh":     str(row.get("MH", "")),
                    "veh":    int(row.get("assigned_vehicle_length", 0) or 0),
                    "dist":   round(dist, 1),
                    "rspkm":  rspkm,
                    "freq":   int(freq),
                    "trips":  int(freq * DAYS_PER_MONTH),
                    "cost":   round(cost, 0),
                })
        return rows

    before_rows = _route_rows([pr.before_from_result, pr.before_to_result])
    after_rows  = _route_rows([pr.after_from_result,  pr.after_to_result])

    # ── Row 1: Section headers ────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value, c.font, c.fill, c.alignment = (
        "CURRENT ROUTES  (before reassignment — Agent 4 ILP)",
        _f(bold=True, sz=10, color="843C0C"), B_FILL, _al("center")
    )
    ws.merge_cells("K1:S1")
    c = ws["K1"]
    c.value, c.font, c.fill, c.alignment = (
        "NEW ROUTES  (after reassignment — Agent 4 ILP)",
        _f(bold=True, sz=10, color="375623"), A_FILL, _al("center")
    )
    ws.row_dimensions[1].height = 18

    # ── Row 2: Column headers ─────────────────────────────────────────────────
    HDRS = ["Route", "DHs on Route", "Serving MH", "Veh (ft)",
            "Dist (km)", "Rs/km", "Freq", "Trips/mo", "Monthly Cost (Rs)"]
    for ci, h in enumerate(HDRS, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = (
            _f(bold=True, sz=9), H_FILL, _al("center"), _border
        )
    for ci, h in enumerate(HDRS, 11):
        c = ws.cell(row=2, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = (
            _f(bold=True, sz=9), H_FILL, _al("center"), _border
        )

    # ── Data rows ─────────────────────────────────────────────────────────────
    def _write_row(rnum, col_start, rdata, rfill):
        vals = [rdata["route"], rdata["dhs"], rdata["mh"], rdata["veh"],
                rdata["dist"], rdata["rspkm"], rdata["freq"], rdata["trips"], rdata["cost"]]
        for ofs, v in enumerate(vals):
            ci = col_start + ofs
            c  = ws.cell(row=rnum, column=ci, value=v)
            c.font, c.fill, c.border = _f(sz=9), rfill, _border
            if ofs == 0:
                c.alignment = _al("left")
            elif ofs == 8:
                c.alignment   = _al("right")
                c.number_format = "#,##0"
            else:
                c.alignment = _al("center")

    n_data = max(len(before_rows), len(after_rows), 0)
    for i in range(n_data):
        rnum  = 3 + i
        rfill = W_FILL if i % 2 == 0 else AL_FILL
        if i < len(before_rows):
            _write_row(rnum, 1,  before_rows[i], rfill)
        if i < len(after_rows):
            _write_row(rnum, 11, after_rows[i],  rfill)

    # ── Totals row ────────────────────────────────────────────────────────────
    tr = 3 + n_data
    b_total = sum(r["cost"] for r in before_rows)
    a_total = sum(r["cost"] for r in after_rows)

    ws.merge_cells(f"A{tr}:H{tr}")
    c = ws.cell(row=tr, column=1, value="TOTAL POOL COST  (Agent 4 ILP — exact)")
    c.font, c.fill, c.alignment = _f(bold=True, sz=9), TO_FILL, _al("right")
    c = ws.cell(row=tr, column=9, value=b_total)
    c.font, c.fill, c.alignment, c.number_format = (
        _f(bold=True, sz=9), TO_FILL, _al("right"), "#,##0"
    )
    ws.merge_cells(f"K{tr}:R{tr}")
    c = ws.cell(row=tr, column=11, value="TOTAL POOL COST  (Agent 4 ILP — exact)")
    c.font, c.fill, c.alignment = _f(bold=True, sz=9), TO_FILL, _al("right")
    c = ws.cell(row=tr, column=19, value=a_total)
    c.font, c.fill, c.alignment, c.number_format = (
        _f(bold=True, sz=9), TO_FILL, _al("right"), "#,##0"
    )

    # ── MH-DH saving row ─────────────────────────────────────────────────────
    sr = tr + 1
    conf = "Confirmed ✓" if pr.savings > SAVING_CONFIRMED_THRESHOLD_RS else "Below threshold"
    saving_text = (
        f"MH-DH SAVING  =  Rs {b_total:,.0f}  –  Rs {a_total:,.0f}"
        f"  =  Rs {pr.savings:,.0f} / month          {conf}"
    )
    ws.merge_cells(f"A{sr}:S{sr}")
    c = ws[f"A{sr}"]
    c.value, c.font, c.fill, c.alignment = (
        saving_text, _f(bold=True, sz=10, color="375623"), SV_FILL, _al("center")
    )
    ws.row_dimensions[sr].height = 18



# ─────────────────────────────────────────────────────────────────────────────
# Phase 7: Write Excel outputs
# ─────────────────────────────────────────────────────────────────────────────

def write_excel_outputs(
    pair_results: list[MHPairResult],
    out_dir:      Path,
) -> list[Path]:
    """Write one Excel per pair and Phase2_Summary.xlsx."""
    out_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []

    import openpyxl as _openpyxl

    for pr in pair_results:
        safe_from = pr.from_mh.replace("/", "_").replace("\\", "_")
        safe_to   = pr.to_mh.replace("/", "_").replace("\\", "_")
        fpath     = out_dir / f"{safe_from}_{safe_to}_phase2.xlsx"

        # Step 1: write data sheets with pandas
        with pd.ExcelWriter(fpath, engine="openpyxl") as xl:
            for sheet_name, df in pr.sheets.items():
                if df is not None and not df.empty:
                    df.to_excel(xl, sheet_name=sheet_name[:31], index=False)

        # Step 2: add formatted Routes sheet and move it to position 0
        try:
            wb = _openpyxl.load_workbook(fpath)
            _write_routes_sheet(wb, pr)
            # Move "Routes" to front
            idx = wb.sheetnames.index("Routes")
            if idx > 0:
                wb.move_sheet("Routes", offset=-idx)
            wb.save(fpath)
        except Exception as _re:
            pass  # formatting failure must not break the data file

        written.append(fpath)
        # Save pool map HTML alongside the Excel
        if pr.map_html:
            map_path = out_dir / f"{safe_from}_{safe_to}_pool_map.html"
            map_path.write_text(pr.map_html, encoding="utf-8")
            written.append(map_path)

    summary_rows = [
        {
            "From_MH":              pr.from_mh,
            "To_MH":                pr.to_mh,
            "Agent3_Flagged_DHs":   len(pr.flagged_dhs),
            "Pool_DHs":             len(pr.pool_dhs),
            "DHs_Moved":            len(pr.dhs_moved),
            "MHDH_Saving_Rs":       round(pr.savings, 2),
            "MHMH_Saving_Rs":       round(pr.mhmh_delta_rs, 2)
                                    if pr.mhmh_delta_rs is not None else None,
            "Total_Savings_Rs":     round(pr.savings + (pr.mhmh_delta_rs or 0.0), 2),
            "Agent3_Est_Saving_Rs": pr.agent3_est_saving_rs,
        }
        for pr in pair_results
    ]
    summary_path = out_dir / "Phase2_Summary.xlsx"
    with pd.ExcelWriter(summary_path, engine="openpyxl") as xl:
        pd.DataFrame(summary_rows).sort_values(
            "Total_Savings_Rs", ascending=False
        ).to_excel(xl, sheet_name="Summary", index=False)
    written.append(summary_path)

    return written


# ─────────────────────────────────────────────────────────────────────────────
# Main pipeline
# ─────────────────────────────────────────────────────────────────────────────

def run_phase2_pipeline(
    agent3_output_path:       Path,
    h2h_path:                 Path,
    location_file_path:       Path,
    lat_long_path:            Path,
    distance_matrix_path:     Path,
    mh_rate_card_path:        Path,
    out_dir:                  Path,
    config_path:              Optional[Path] = None,
    residual_threshold:       float = 100.0,
    min_pair_savings_rs:      float = 0.0,
    target_pairs:             Optional[list[tuple[str, str]]] = None,
    on_progress:              Optional[Callable[[str], None]] = None,
    plan_vol_path:            Optional[Path] = None,
    mh1_mh2_cost_matrix_path: Optional[Path] = None,
) -> Phase2Result:
    """
    Full Phase 2 pipeline.

    1.  Load all inputs
    2.  Compute MH-pair savings from Agent 3 output
    3.  For each qualifying pair:
        a.  Expand DH pool via H2H MR-group logic
        b.  Batch gap-fill MH-MH costs for route-mates / non-candidate target MHs
            (calls compute_mhmh_for_pairs when plan_vol_path + mh1_mh2_cost_matrix_path
             are available and Agent 3 candidate columns don't cover the target MH)
        c.  Log per-DH MH-MH source (Agent 3 column / gap-fill / unavailable)
        d.  Precompute non-pool cost once (fixed offset — Issue 4)
        e.  Optimise assignment using full Agent 4 ILP (Issues 1/5)
        f.  Run Agent 4 before + after on pool DHs to get true costs
            (baseline uses current_fc_mh assignments — Issue 6 confirmed correct)
        g.  Build 5-sheet Excel content (Issues 2/3/7/I11)
    4.  Write Excel outputs
    5.  Print console summary
    6.  Return Phase2Result

    plan_vol_path / mh1_mh2_cost_matrix_path
        Optional — supplied by orchestrator via paths_for_agent3_phase2_run().
        When provided, Phase 2 batch-computes exact MH-MH costs for any pool DHs
        whose target MH is not covered by Agent 3's top-4 candidates.
        If omitted, MH-MH is "unavailable" for those DHs — no proxy approximations.
    """

    def _prog(msg: str) -> None:
        if on_progress:
            on_progress(msg)

    _prog("Loading inputs ...")

    agent3_df = pd.read_csv(agent3_output_path, dtype=str)

    h2h_df = load_h2h(h2h_path)

    location_df = pd.read_excel(location_file_path, dtype=str)
    location_df.columns = location_df.columns.str.strip()
    for col in ("total_cft", "top266_shipments", "ML",
                "time_window_start (minutes)",
                "time_window_end (minutes)",
                "depot_departure (minutes)"):
        if col in location_df.columns:
            location_df[col] = pd.to_numeric(location_df[col], errors="coerce")

    ll_df = pd.read_excel(lat_long_path, dtype=str)
    ll_df.columns = ll_df.columns.str.strip()
    latlong: dict[str, tuple[float, float]] = {}
    for _, row in ll_df.iterrows():
        name = str(row.get("Site_name", "")).strip()
        try:
            latlong[name] = (float(row["Latitude"]), float(row["Longitude"]))
        except (ValueError, TypeError, KeyError):
            pass

    dist_df = pd.read_csv(distance_matrix_path, dtype=str)
    dist_df.columns = dist_df.columns.str.strip()
    dist_dict: dict[tuple[str, str], float] = {}
    for _, row in dist_df.iterrows():
        src = str(row.get("S_Code", "")).strip()
        dst = str(row.get("D_Code", "")).strip()
        try:
            dist_dict[(src, dst)] = float(row["distance"])
        except (ValueError, TypeError, KeyError):
            pass

    cfg        = p4.load_agent4_config(config_path)
    mh_configs = p4.load_rate_card(mh_rate_card_path, cfg)

    _prog("Computing MH-pair savings ...")
    pair_savings_df = compute_mh_pair_savings(agent3_df)
    qualifying = pair_savings_df[
        pair_savings_df["total_savings_rs"] >= min_pair_savings_rs
    ]

    # Optional explicit pair filter — list of (from_mh, to_mh) tuples
    if target_pairs:
        target_set = {(a.strip(), b.strip()) for a, b in target_pairs}
        qualifying = qualifying[
            qualifying.apply(
                lambda r: (str(r["from_mh"]).strip(), str(r["to_mh"]).strip()) in target_set,
                axis=1,
            )
        ].reset_index(drop=True)

    pair_results: list[MHPairResult] = []
    n_pairs = len(qualifying)

    # ── Pre-flight: expand all pools then print summary table ─────────────────
    _prog("")
    _prog("Pre-flight: expanding pools and determining optimisation mode ...")
    preflight: list[dict] = []
    for _, row in qualifying.iterrows():
        fm  = str(row["from_mh"])
        tm  = str(row["to_mh"])
        flg = list(row["dhs"])
        pl  = expand_pool(flg, fm, tm, h2h_df, agent3_df, location_df)
        n   = len(pl)
        if n <= FULL_ENUM_THRESHOLD:
            mode    = "Full enum"
            a4_runs = (1 << n) * 2 + 2          # 2^n scenarios × 2 + baseline
        else:
            mode    = f"LNS ({LNS_ITERATIONS} iter)"
            a4_runs = LNS_ITERATIONS * 2 + 2    # 500 iter × 2 + baseline
        preflight.append({"from_mh": fm, "to_mh": tm,
                           "flagged": flg,       # original list, not just count
                           "pool": pl,
                           "mode": mode, "a4_runs": a4_runs})

    # Short MH labels for display (strip CENTRALHUB_L_ prefix)
    def _short(mh: str) -> str:
        for pre in ("CENTRALHUB_L_", "CENTRALHUB_"):
            if mh.upper().startswith(pre.upper()):
                return mh[len(pre):]
        return mh

    W_PAIR, W_POOL, W_MODE, W_RUNS = 32, 10, 20, 14
    divider = "-" * (W_PAIR + W_POOL + W_MODE + W_RUNS + 3)
    header  = (f"{'Pair':<{W_PAIR}} {'Pool DHs':>{W_POOL}}"
               f"  {'Mode':<{W_MODE}} {'Agent 4 runs':>{W_RUNS}}")
    _prog("")
    _prog("OPTIMISATION PLAN")
    _prog(divider)
    _prog(header)
    _prog(divider)
    for pf in preflight:
        pair_label = f"{_short(pf['from_mh'])} -> {_short(pf['to_mh'])}"
        flag       = "   OK" if pf["mode"].startswith("Full") else "   WARN: heuristic"
        _prog(f"{pair_label:<{W_PAIR}} {len(pf['pool']):>{W_POOL},}"
              f"  {pf['mode']:<{W_MODE}} {pf['a4_runs']:>{W_RUNS},}{flag}")
    _prog(divider)
    lns_pairs = [pf for pf in preflight if not pf["mode"].startswith("Full")]
    if lns_pairs:
        _prog(f"WARNING: {len(lns_pairs)} pair(s) will use LNS (heuristic) — "
              f"pool > {FULL_ENUM_THRESHOLD} DHs. Result may not be globally optimal.")
    _prog("")

    for enum_idx, pf in enumerate(preflight):
        from_mh      = pf["from_mh"]
        to_mh        = pf["to_mh"]
        flagged_orig = pf["flagged"]   # original Agent-3-flagged DHs list
        pool         = pf["pool"]      # H2H-expanded pool (already computed)

        _prog(f"[{enum_idx + 1}/{n_pairs}] {from_mh} -> {to_mh}  "
              f"({len(flagged_orig)} flagged | pool={len(pool)} | "
              f"{pf['mode']} | ~{pf['a4_runs']:,} Agent 4 runs)")
        _prog(f"  Pool: {len(flagged_orig)} flagged -> {len(pool)} after H2H expansion")

        pool_set = set(pool)

        # ── Batch gap-fill: MH-MH costs for DHs whose target MH was not in ──
        # Agent 3's top-4 candidates (route-mates, non-candidate target MH).
        # Uses compute_mhmh_for_pairs when orchestrator paths are available.
        mhmh_cache: dict[tuple[str, str], Optional[float]] = {}
        if plan_vol_path and mh1_mh2_cost_matrix_path:
            need_gap_fill: list[tuple[str, str]] = []
            for dh in pool:
                r = agent3_df[agent3_df["destination_hub_key"] == dh]
                if r.empty:
                    need_gap_fill.append((dh, to_mh))
                    continue
                rv = r.iloc[0]
                found = False
                for i in range(1, 5):
                    cand_mh = str(rv.get(f"candidate_{i}", "") or "").strip()
                    if cand_mh.upper() == to_mh.strip().upper():
                        mhmh_val = pd.to_numeric(
                            rv.get(f"candidate_{i}_mhmh_cost_rs"), errors="coerce"
                        )
                        if pd.notna(mhmh_val):
                            found = True
                            break
                if not found:
                    need_gap_fill.append((dh, to_mh))

            if need_gap_fill:
                _prog(f"  Gap-fill: computing MH-MH for {len(need_gap_fill)} DH(s) "
                      f"without candidate_X_mhmh_cost_rs data ...")
                try:
                    gap_results = _a3_compute_mhmh(
                        need_gap_fill,
                        plan_vol_path=plan_vol_path,
                        cost_matrix_path=mh1_mh2_cost_matrix_path,
                        config_path=config_path,
                    )
                    mhmh_cache.update(gap_results)
                    filled = sum(1 for v in gap_results.values() if v is not None)
                    _prog(f"  Gap-fill done: {filled}/{len(need_gap_fill)} resolved")
                except Exception as _gfe:
                    _prog(f"  Gap-fill WARNING: {_gfe} — MH-MH will be unavailable for those DHs")

        # ── Per-DH MH-MH source log ───────────────────────────────────────────
        _prog(f"  MH-MH cost source for {len(pool)} pool DH(s) → {to_mh}:")
        for _dh in pool:
            _r = agent3_df[agent3_df["destination_hub_key"] == _dh]
            _src = None
            if not _r.empty:
                _rv = _r.iloc[0]
                for _i in range(1, 5):
                    _cmh = str(_rv.get(f"candidate_{_i}", "") or "").strip()
                    if _cmh.upper() == to_mh.strip().upper():
                        _val = pd.to_numeric(
                            _rv.get(f"candidate_{_i}_mhmh_cost_rs"), errors="coerce"
                        )
                        if pd.notna(_val):
                            _src = f"Agent 3 candidate_{_i}_mhmh_cost_rs  Rs {_val:,.0f}"
                            break
            if _src is None:
                _cached = mhmh_cache.get((_dh, to_mh))
                if _cached is not None:
                    _src = f"gap-fill (on-demand)  Rs {_cached:,.0f}"
                else:
                    _src = "UNAVAILABLE"
            _prog(f"    {_dh}: {_src}")

        # Build before / initial assignments from Agent 3 data.
        # If Agent 3 recommended a third MH (neither from_mh nor to_mh), clamp
        # to from_mh so Phase 2 evaluates a clean stay-vs-migrate decision for
        # this pair.  The third-MH recommendation is handled by its own Phase 2
        # pair (e.g. MPL1→HBL1) where that DH is a proper flagged DH.
        a3_assign:     dict[str, str] = {}
        before_assign: dict[str, str] = {}
        for dh in pool:
            r = agent3_df[agent3_df["destination_hub_key"] == dh]
            if not r.empty:
                _a3_rec = str(r.iloc[0].get("assigned_fc_mh", from_mh))
                a3_assign[dh]     = _a3_rec if _a3_rec in (from_mh, to_mh) else from_mh
                before_assign[dh] = str(r.iloc[0].get("current_fc_mh",  from_mh))
            else:
                a3_assign[dh]     = from_mh
                before_assign[dh] = from_mh

        # ── Pre-compute MHMH monthly costs for both MHs ──────────────────────
        # These are passed into the optimizer so every iteration (base, all
        # enumeration combos, LNS repair + evaluate) scores MHDH + MHMH.
        # DHs where the cost is unavailable default to 0.0 — no phantom
        # penalty or reward for missing data.
        _DAYS = DAYS_PER_MONTH
        mhmh_monthly_mh1: dict[str, float] = {}
        mhmh_monthly_mh2: dict[str, float] = {}
        for dh in pool:
            _c1, _ = _mhmh_cost_at_mh(dh, from_mh, agent3_df, mhmh_cache)
            _c2, _ = _mhmh_cost_at_mh(dh, to_mh,   agent3_df, mhmh_cache)
            mhmh_monthly_mh1[dh] = float(_c1) * _DAYS if _c1 is not None else 0.0
            mhmh_monthly_mh2[dh] = float(_c2) * _DAYS if _c2 is not None else 0.0
        _n_m1 = sum(1 for v in mhmh_monthly_mh1.values() if v > 0)
        _n_m2 = sum(1 for v in mhmh_monthly_mh2.values() if v > 0)
        _prog(f"  MHMH pre-compute: {_n_m1}/{len(pool)} resolved at {from_mh}, "
              f"{_n_m2}/{len(pool)} resolved at {to_mh}")

        # ── Phase 5: Optimise (MHDH + MHMH joint scoring) ────────────────────
        _prog(f"  Optimising {len(pool)} pool DHs (MHDH + MHMH end-to-end scoring) ...")
        best, best_afr, best_atr = optimize_pool_assignment(
            pool_dhs=pool,
            mh1=from_mh, mh2=to_mh,
            dist_dict=dist_dict, latlong=latlong,
            mh_configs=mh_configs, location_df=location_df,
            cfg=cfg, residual_threshold=residual_threshold,
            initial_assignment=a3_assign,
            on_progress=_prog,
            mhmh_monthly_mh1=mhmh_monthly_mh1,
            mhmh_monthly_mh2=mhmh_monthly_mh2,
        )

        # ── Baseline (before) costs — Issue 6 confirmed: uses run_agent4_for_mh
        from_before = [d for d in pool if before_assign.get(d, from_mh) == from_mh]
        to_before   = [d for d in pool if before_assign.get(d, from_mh) == to_mh]
        _prog(f"  Running Agent 4 (before): {from_mh}={len(from_before)}, "
              f"{to_mh}={len(to_before)}")
        bfr = _run_a4_subset(from_mh, from_before, dist_dict, latlong,
                              mh_configs, location_df, cfg, residual_threshold)
        btr = _run_a4_subset(to_mh,   to_before,   dist_dict, latlong,
                              mh_configs, location_df, cfg, residual_threshold)

        # best_afr / best_atr already computed inside optimize_pool_assignment
        # Use them directly rather than re-running Agent 4
        afr, atr = best_afr, best_atr

        pool_cost_before = bfr.total_monthly_cost + btr.total_monthly_cost
        pool_cost_after  = afr.total_monthly_cost + atr.total_monthly_cost
        cost_before      = pool_cost_before
        cost_after       = pool_cost_after
        savings          = cost_before - cost_after

        dhs_moved  = [d for d in pool if best.get(d) == to_mh]
        dhs_stayed = [d for d in pool if best.get(d, from_mh) == from_mh]

        # Phase 6: build 5-sheet output
        sheets, pair_mhmh_delta, pair_a3_est_saving = build_pair_output(
            from_mh=from_mh, to_mh=to_mh,
            pool_dhs=pool,
            best_assignment=best,
            agent3_df=agent3_df,
            location_df=location_df,
            h2h_df=h2h_df,
            dist_dict=dist_dict,
            mh_configs=mh_configs,
            cfg=cfg,
            before_from_result=bfr, before_to_result=btr,
            after_from_result=afr,  after_to_result=atr,
            mhmh_cache=mhmh_cache if mhmh_cache else None,
            flagged_dhs=flagged_orig,
        )

        # Build pool map (mirrors Agent 3 style — pool DHs only)
        try:
            map_html = _build_phase2_map_html(
                pr=MHPairResult(  # temporary lightweight object just for map builder
                    from_mh=from_mh, to_mh=to_mh,
                    flagged_dhs=flagged_orig, pool_dhs=pool,
                    best_assignment=best,
                    pool_cost_before=pool_cost_before,
                    pool_cost_after=pool_cost_after,
                    cost_before=cost_before,
                    cost_after=cost_after,
                    savings=savings,
                    mhmh_delta_rs=pair_mhmh_delta,
                    dhs_moved=dhs_moved, dhs_stayed=dhs_stayed,
                ),
                latlong=latlong,
                location_df=location_df,
                per_dh_detail_df=sheets.get("Per_DH_Detail"),
            )
        except Exception as _me:
            _prog(f"  Map build WARNING: {_me}")
            map_html = None

        total_savings_rs = savings + (pair_mhmh_delta or 0.0)
        pair_results.append(MHPairResult(
            from_mh=from_mh, to_mh=to_mh,
            flagged_dhs=flagged_orig, pool_dhs=pool,
            best_assignment=best,
            pool_cost_before=pool_cost_before,
            pool_cost_after=pool_cost_after,
            cost_before=cost_before,
            cost_after=cost_after,
            savings=savings,
            mhmh_delta_rs=pair_mhmh_delta,
            agent3_est_saving_rs=pair_a3_est_saving,
            dhs_moved=dhs_moved, dhs_stayed=dhs_stayed,
            before_from_result=bfr, before_to_result=btr,
            after_from_result=afr,  after_to_result=atr,
            sheets=sheets,
            map_html=map_html,
        ))
        mhmh_str = (f"  MHMH delta: Rs {pair_mhmh_delta:,.0f}/month"
                    if pair_mhmh_delta is not None else "  MHMH delta: unavailable")
        _prog(f"  DONE  MHDH saving: Rs {savings:,.0f}/month  |{mhmh_str}  "
              f"|  Total: Rs {total_savings_rs:,.0f}/month  "
              f"({len(dhs_moved)} DHs moved, {len(dhs_stayed)} stayed)")

    # Phase 7: write Excel
    _prog("Writing Excel outputs ...")
    excel_paths = write_excel_outputs(pair_results, out_dir)

    # ── Console summary ───────────────────────────────────────────────────────
    summary_rows = [
        {
            "From_MH":              pr.from_mh,
            "To_MH":                pr.to_mh,
            "Agent3_Flagged_DHs":   len(pr.flagged_dhs),
            "Pool_DHs":             len(pr.pool_dhs),
            "DHs_Moved":            len(pr.dhs_moved),
            "MHDH_Saving_Rs":       round(pr.savings, 2),
            "MHMH_Saving_Rs":       round(pr.mhmh_delta_rs, 2)
                                    if pr.mhmh_delta_rs is not None else None,
            "Total_Savings_Rs":     round(pr.savings + (pr.mhmh_delta_rs or 0.0), 2),
            "Agent3_Est_Saving_Rs": pr.agent3_est_saving_rs,
        }
        for pr in pair_results
    ]
    summary_df = pd.DataFrame(summary_rows)

    if not summary_df.empty:
        _prog("\n" + "=" * 70)
        _prog("PHASE 2 SUMMARY")
        _prog("=" * 70)
        for _, sr in summary_df.iterrows():
            mhmh_str = (f"Rs {sr['MHMH_Saving_Rs']:,.0f}"
                        if sr["MHMH_Saving_Rs"] is not None else "N/A")
            a3_str   = (f"Rs {sr['Agent3_Est_Saving_Rs']:,.0f}"
                        if sr["Agent3_Est_Saving_Rs"] is not None else "N/A")
            _prog(
                f"  {sr['From_MH']} -> {sr['To_MH']}"
                f"  |  flagged={sr['Agent3_Flagged_DHs']} pool={sr['Pool_DHs']} moved={sr['DHs_Moved']}"
                f"  |  MHDH=Rs {sr['MHDH_Saving_Rs']:,.0f}"
                f"  |  MHMH={mhmh_str}"
                f"  |  total=Rs {sr['Total_Savings_Rs']:,.0f}"
                f"  |  A3 est={a3_str}"
            )
        total_saving = summary_df["Total_Savings_Rs"].sum()
        _prog(f"\n  Grand total saving: Rs {total_saving:,.0f}/month")
        _prog("=" * 70)

    return Phase2Result(
        pair_results=pair_results,
        summary_df=summary_df,
        excel_paths=excel_paths,
    )
